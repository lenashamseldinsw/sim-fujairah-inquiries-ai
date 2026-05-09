"""
STAGE 6: Artifact Generator

Generates:
1. Excel workbook (openpyxl) — 10 sheets with classification results
2. Word report (sword-word-builder) — 9-section bilingual report
3. Report dictionary — demo-compatible format for Streamlit display (in-memory)

Uses state.report_sections_ar and state.report_sections_en to build Word document and report dict.
Report dict is stored in state.report_json for passing to display functions.
"""

import json
import sys
import anthropic
from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import importlib.util
from concurrent.futures import ThreadPoolExecutor, as_completed

from .state import PipelineState, CaseRow, convert_month_year_to_arabic
from .stage6_json_report import generate_json_report
from .generate_workload_map_section import generate_workload_map_section
from .generate_customer_journey_section import generate_customer_journey_section
from .generate_digital_gaps_section import generate_digital_gaps_section
from .generate_digital_transformation_section import generate_digital_transformation_section
from .generate_ai_use_cases_section import generate_ai_use_cases_section
from .generate_improvement_roadmap_section import generate_improvement_roadmap_section
from .generate_conclusion_section import generate_conclusion_section

# Import build_report_ar to generate Word document from JSON
try:
    from .build_report_ar import build_report
except ImportError:
    build_report = None
    print("[Warning] Could not import build_report_ar")


# Excel formatting
HEADER_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_ROW_FILL = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Sheet names in order
SHEET_NAMES = [
    'ملخص',                    # Summary
    'كل الحالات',              # All Cases
    'طلبات',                   # Service Requests
    'استفسارات',               # Information Inquiries
    'متابعات',                 # Status Follow-ups
    'مشاكل جهات أخرى',         # Cross-Entity
    'استفسارات الموقع',        # Location Inquiries
    'بلاغات تقنية',            # Tech Incidents
    'استفسارات مالية',         # Financial Inquiries
    'إعادة التصنيف',           # Misclassified Cases
]


def generate_excel(state: PipelineState, output_path: str) -> None:
    """Generate Excel workbook with classification results."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Summary
    ws_summary = wb.create_sheet('ملخص', 0)
    _populate_summary_sheet(ws_summary, state)

    # Sheet 2: All Cases
    ws_all = wb.create_sheet('كل الحالات', 1)
    _populate_all_cases_sheet(ws_all, state.all_classified, state)

    # Sheets 3–6: One per top-level type, filtered and populated
    type_sheet_map = [
        ('شكاوى',      'شكوى'),
        ('طلبات',      'طلب'),
        ('استفسارات',  'استفسار'),
        ('شكر وثناء',  'شكر وثناء'),
    ]
    for idx, (sheet_name, type_value) in enumerate(type_sheet_map, 2):
        subset = [c for c in state.all_classified if c.actual_contact_type == type_value]
        ws_type = wb.create_sheet(sheet_name, idx)
        _populate_all_cases_sheet(ws_type, subset, state)

    # Sheet 7: Reclassified cases
    ws_misclass = wb.create_sheet('إعادة التصنيف', 6)
    misclassified = [c for c in state.all_classified if c.misclassification != 'OK']
    _populate_all_cases_sheet(ws_misclass, misclassified, state)
    ws_misclass.sheet_properties.tabColor = "C0392B"

    wb.save(output_path)


def _populate_summary_sheet(ws, state: PipelineState) -> None:
    """Populate summary sheet with all 5 sections matching sample output."""
    # Enable RTL layout
    ws.sheet_view.rightToLeft = True

    total = len(state.all_classified)

    ws.merge_cells('A1:D1')
    ws['A1'] = f"ملخص تحليل استفسارات شرطة الفجيرة — {convert_month_year_to_arabic(state.month_year) or 'Q1 2026'}"
    ws['A1'].font = Font(bold=True, size=14)

    ws.merge_cells('A2:D2')
    ws['A2'] = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d')}"

    # --- Section 1: Total cases ---
    row = 4
    ws[f'A{row}'] = "1. إجمالي الحالات"
    ws[f'A{row}'].font = Font(bold=True)
    row = 5
    ws[f'A{row}'] = "إجمالي الحالات المعالجة"
    ws[f'B{row}'] = total
    ws[f'C{row}'] = "100.0%"

    # --- Section 2: Distribution by classification type ---
    row = 7
    ws[f'A{row}'] = "2. توزيع حسب نوع التصنيف"
    ws[f'A{row}'].font = Font(bold=True)
    type_counts = {}
    for case in state.all_classified:
        ct = case.actual_contact_type
        type_counts[ct] = type_counts.get(ct, 0) + 1
    row = 8
    for ct, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
        ws[f'A{row}'] = ct
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"{(count / total * 100):.1f}%" if total else "0.0%"
        row += 1

    # --- Section 3: Distribution by channel ---
    row += 1
    ws[f'A{row}'] = "3. توزيع حسب قناة التقديم"
    ws[f'A{row}'].font = Font(bold=True)
    channel_counts = {}
    for case in state.all_classified:
        ch = (case.case_channel or '').strip() or 'غير محدد'
        channel_counts[ch] = channel_counts.get(ch, 0) + 1
    row += 1
    for ch, count in sorted(channel_counts.items(), key=lambda x: x[1], reverse=True):
        ws[f'A{row}'] = ch
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"{(count / total * 100):.1f}%" if total else "0.0%"
        row += 1

    # --- Section 4: SLA on-time rate ---
    row += 1
    ws[f'A{row}'] = "4. معدل الإغلاق في الوقت المحدد"
    ws[f'A{row}'].font = Font(bold=True)
    on_time = sum(1 for c in state.all_classified if str(c.sla_color).strip() == 'نعم')
    late = total - on_time
    row += 1
    ws[f'A{row}'] = "تم الإغلاق في الوقت"
    ws[f'B{row}'] = on_time
    ws[f'C{row}'] = f"{(on_time / total * 100):.1f}%" if total else "0.0%"
    row += 1
    ws[f'A{row}'] = "تجاوز الوقت المحدد"
    ws[f'B{row}'] = late
    ws[f'C{row}'] = f"{(late / total * 100):.1f}%" if total else "0.0%"

    # --- Section 5: Reclassification count ---
    row += 2
    reclassified_count = state.reclassified_count
    matched_count = total - reclassified_count
    ws[f'A{row}'] = f"5. إعادة التصنيف ({reclassified_count} حالة)"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "حالات أُعيد تصنيفها"
    ws[f'B{row}'] = reclassified_count
    ws[f'C{row}'] = f"{(reclassified_count / total * 100):.1f}%" if total else "0.0%"
    row += 1
    ws[f'A{row}'] = "حالات تطابقت مع التصنيف الأصلي"
    ws[f'B{row}'] = matched_count
    ws[f'C{row}'] = f"{(matched_count / total * 100):.1f}%" if total else "0.0%"

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12


def _populate_all_cases_sheet(ws, cases: List[CaseRow], state: PipelineState) -> None:
    """Populate sheet with case data (RTL).

    Uses fixed column order across all sheets:
    رقم_الطلب, تفاصيل_الطلب, الحل, الخدمة, الخدمة_الرئيسية, نوع_المكالمة,
    التصنيف_الفعلي, التصنيف_الفرعي, السبب, إعادة_التصنيف, قناة_تقديم_الخدمة,
    الحالة_SLA, تاريخ_الإنشاء, الإدارة_العامة
    """
    # Enable RTL layout
    ws.sheet_view.rightToLeft = True

    # Map normalized column names to CaseRow attributes
    NORMALIZED_TO_CASEROW = {
        'رقم_الطلب': 'case_number',
        'تفاصيل_الطلب': 'case_title',
        'تاريخ_الإنشاء': 'date_opened',
        'قناة_تقديم_الخدمة': 'case_channel',
        'نوع_المكالمة': 'case_type',
        'الخدمة_الرئيسية': 'service_name',
        'الحل': 'resolution_response',
        'الحالة_SLA': 'sla_color',
        'الإدارة_العامة': 'admin',
        'الخدمة': 'service',  # May not exist in CaseRow — pull from raw_df if available
    }

    # Import COLUMN_MAPPING for validation
    from .stage1_validator import COLUMN_MAPPING

    # FIXED COLUMN ORDER (all data sheets use this exact sequence)
    # Input columns (1-6), then AI-generated (7-10), then remaining input (11-14)
    headers = [
        'رقم_الطلب',               # 1
        'تفاصيل_الطلب',            # 2
        'الحل',                    # 3
        'الخدمة',                  # 4
        'الخدمة_الرئيسية',         # 5
        'نوع_المكالمة',            # 6
        'التصنيف_الفعلي',          # 7 (AI-generated)
        'التصنيف_الفرعي',          # 8 (AI-generated)
        'السبب',                   # 9 (AI-generated)
        'إعادة_التصنيف',           # 10 (AI-generated)
        'قناة_تقديم_الخدمة',        # 11
        'الحالة_SLA',              # 12
        'تاريخ_الإنشاء',           # 13
        'الإدارة_العامة',          # 14
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Create a lookup dict for raw_df by case_number for columns not in CaseRow
    raw_df_lookup = {}
    if state.raw_df is not None:
        try:
            import pandas as pd
            df = state.raw_df
            # Get the normalized case number column
            case_num_col = 'رقم_الطلب'
            if case_num_col in df.columns:
                # Create lookup by case number
                for idx, row in df.iterrows():
                    case_num = str(row[case_num_col])
                    raw_df_lookup[case_num] = row
        except Exception as e:
            print(f"[Warning] Could not create raw_df lookup: {e}")

    # Write data rows
    for row_idx, case in enumerate(cases, 2):
        # Convert misclassification to نعم/لا
        reclassified = 'لا' if case.misclassification == 'OK' else 'نعم'

        # Get raw row data for this case (for unmapped columns)
        raw_row = raw_df_lookup.get(case.case_number)

        # Build data in the exact order specified by headers
        # Note: Headers are interleaved with input + AI columns
        row_data = []

        for header_col in headers:
            value = ''

            # Input columns from CaseRow
            if header_col == 'رقم_الطلب':
                value = case.case_number
            elif header_col == 'تفاصيل_الطلب':
                value = case.case_title or ''
            elif header_col == 'الحل':
                value = case.resolution_response or ''
            elif header_col == 'الخدمة':
                # Try to get من raw_df (not in CaseRow)
                if raw_row is not None:
                    for col_name in ['الخدمة', 'الخدمة ', 'Service']:
                        if col_name in raw_row.index:
                            val = raw_row[col_name]
                            if val == val:  # pd.notna check
                                value = val
                            break
            elif header_col == 'الخدمة_الرئيسية':
                value = case.service_name or ''
            elif header_col == 'نوع_المكالمة':
                value = case.case_type or ''
            elif header_col == 'قناة_تقديم_الخدمة':
                value = case.case_channel or ''
            elif header_col == 'الحالة_SLA':
                value = case.sla_color or ''
            elif header_col == 'تاريخ_الإنشاء':
                value = case.date_opened or ''
            elif header_col == 'الإدارة_العامة':
                value = case.admin or ''
            # AI-generated columns
            elif header_col == 'التصنيف_الفعلي':
                value = case.actual_contact_type or ''
            elif header_col == 'التصنيف_الفرعي':
                value = case.sub_classification or ''
            elif header_col == 'السبب':
                value = case.classification_reason or ''
            elif header_col == 'إعادة_التصنيف':
                value = reclassified

            row_data.append(value)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
            cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)

    # Auto-size columns (fixed order per headers)
    ws.column_dimensions['A'].width = 12   # رقم_الطلب
    ws.column_dimensions['B'].width = 80   # تفاصيل_الطلب
    ws.column_dimensions['C'].width = 60   # الحل
    ws.column_dimensions['D'].width = 20   # الخدمة
    ws.column_dimensions['E'].width = 20   # الخدمة_الرئيسية
    ws.column_dimensions['F'].width = 18   # نوع_المكالمة
    ws.column_dimensions['G'].width = 18   # التصنيف_الفعلي
    ws.column_dimensions['H'].width = 18   # التصنيف_الفرعي
    ws.column_dimensions['I'].width = 18   # السبب
    ws.column_dimensions['J'].width = 18   # إعادة_التصنيف
    ws.column_dimensions['K'].width = 18   # قناة_تقديم_الخدمة
    ws.column_dimensions['L'].width = 18   # الحالة_SLA
    ws.column_dimensions['M'].width = 18   # تاريخ_الإنشاء
    ws.column_dimensions['N'].width = 18   # الإدارة_العامة

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(cases) + 1}'


def generate_word_report(
    state: PipelineState,
    output_path: str,
    language: str = 'ar',
    api_key: str = ""
) -> None:
    """
    Generate Word report using build_report_ar.py.

    Generates JSON report, saves it to disk, then uses build_report_ar.build_report()
    to create the styled Word document. Note: build_report_ar only supports Arabic.

    Args:
        state: Pipeline state with report_sections
        output_path: Path to save .docx
        language: Kept for backward compatibility (build_report_ar is Arabic-only)
        api_key: Anthropic API key for LLM report generation
    """
    if build_report is None:
        print(f"⚠️  Skipping Word report generation (build_report_ar not available)")
        return

    if output_path is None:
        print(f"⚠️  Skipping Word report generation (no output_path provided)")
        return

    # Generate JSON report from state
    report_data = generate_json_report(state)

    # Create output path and derive JSON path from it
    output_path = Path(output_path)
    json_path = output_path.with_stem(output_path.stem + "_data").with_suffix(".json")

    # Save JSON to disk
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(report_data, f, ensure_ascii=False, indent=2)
    print(f"[Stage6] Saved report JSON: {json_path}")

    # Build Word document from JSON
    build_report(json_path, output_path)
    print(f"[Stage6] Generated Word report: {output_path}")


def _generate_report_sections(state: PipelineState, api_key: str = "") -> None:
    """
    Generate report sections via LLM with threading.

    Uses 3-wave parallelization:
    - Wave 1: 7 independent sections in parallel (executive_summary, methodology, workload_map,
              customer_journey, digital_gaps, digital_transformation, ai_use_cases)
    - Wave 2: improvement_roadmap (depends on ai_use_cases from wave 1)
    - Wave 3: conclusion (depends on both ai_use_cases and improvement_roadmap from earlier waves)

    Thread safety: Each generator function is read-only on state and returns a result dict.
    Results are collected in a local dict per thread, then merged into state.report_sections_ar
    after all threads complete (single-threaded merge, no race conditions).
    """
    print(f"[GenSections] api_key present: {bool(api_key)}")
    print(f"[GenSections] api_key length: {len(api_key) if api_key else 0}")

    if not api_key:
        raise ValueError("API key is required to generate report sections")

    # Initialize Arabic-only dict
    state.report_sections_ar = {}

    # ──────────────────────────────────────────────────────────────────────────────────
    # WAVE 1: 7 independent sections — run in parallel
    # ──────────────────────────────────────────────────────────────────────────────────
    print("[GenSections] WAVE 1: Starting 7 parallel section generations...")

    wave1_tasks = {
        'executive_summary': (generate_executive_summary_section, 'أولاً: الملخص التنفيذي — التحليلات الرئيسية'),
        'methodology': (generate_methodology_section, 'ثانياً: المنهجية وطبيعة المصادر'),
        'workload_map': (generate_workload_map_section, 'ثالثاً: التحليل الأول — خريطة تصنيف الطلبات'),
        'customer_journey': (generate_customer_journey_section, 'رابعاً: التحليل الثاني — التحديات في رحلة المتعامل'),
        'digital_gaps': (generate_digital_gaps_section, 'خامساً: التحليل الثالث — تحليل الفجوات الرقمية'),
        'digital_transformation': (generate_digital_transformation_section, 'سادساً: التحليل الرابع — خطة التحويل الرقمي'),
        'ai_use_cases': (generate_ai_use_cases_section, 'سابعاً: حالات الاستخدام المدعومة بالذكاء الاصطناعي'),
    }

    wave1_results = {}  # Local dict — no shared state touched inside threads

    def run_section(key: str, fn, state: PipelineState, api_key: str) -> tuple[str, Dict[str, Any]]:
        """Run a section generator with up to 3 retries on failure."""
        max_attempts = 3
        last_error: Exception = RuntimeError(f"[GenSections] {key}: no attempts made")
        for attempt in range(1, max_attempts + 1):
            try:
                if attempt > 1:
                    print(f"[GenSections] Retrying {key} (attempt {attempt}/{max_attempts})...")
                else:
                    print(f"[GenSections] Starting {key}...")
                result = fn(state, api_key)  # Pure: reads state, returns dict
                print(f"[GenSections] ✓ {key} complete (attempt {attempt})")
                return key, result
            except Exception as e:
                last_error = e
                print(f"[GenSections] ✗ {key} failed on attempt {attempt}: {e}")
        raise last_error

    with ThreadPoolExecutor(max_workers=7) as executor:
        futures = {
            executor.submit(run_section, key, fn, state, api_key): key
            for key, (fn, _heading) in wave1_tasks.items()
        }
        for future in as_completed(futures):
            key = futures[future]
            try:
                result_key, result = future.result()
                wave1_results[result_key] = result
            except Exception as e:
                raise RuntimeError(f"[GenSections] Wave 1 failed on '{key}': {e}") from e

    print(f"[GenSections] ✓ WAVE 1 complete: {len(wave1_results)} sections generated")

    # ──────────────────────────────────────────────────────────────────────────────────
    # Merge Wave 1 results into state (single-threaded, no race conditions)
    # ──────────────────────────────────────────────────────────────────────────────────
    print("[GenSections] Merging WAVE 1 results into state...")

    # Special handling for executive_summary and methodology (have special structure)
    if 'executive_summary' in wave1_results:
        exec_summary = wave1_results['executive_summary']
        state.report_sections_ar['executive_summary'] = {
            'heading': wave1_tasks['executive_summary'][1],
            'body': exec_summary.get('framing_paragraph', ''),
            'tables': [exec_summary.get('key_findings', [])],
            'core_message': exec_summary.get('core_message', ''),
            'raw_data': exec_summary
        }

    if 'methodology' in wave1_results:
        methodology = wave1_results['methodology']
        sources_rows = methodology.get('sources_table', [])
        sources_ar = {
            'columns': ['المصدر', 'الطبيعة', 'الحجم', 'الفترة'],
            'rows': sources_rows,
            'row_count': len(sources_rows),
            'col_count': 4,
        }
        if not _is_valid_table(sources_ar):
            raise RuntimeError("[GenSections] Arabic sources table is invalid or empty")
        state.report_sections_ar['methodology'] = {
            'heading': wave1_tasks['methodology'][1],
            'classification_method': methodology.get('classification_method', ''),
            'analyzed_fields': methodology.get('analyzed_fields', ''),
            'tables': [sources_ar],
            'raw_data': methodology
        }

    # Standard sections (workload_map, customer_journey, digital_gaps, digital_transformation, ai_use_cases)
    standard_sections = ['workload_map', 'customer_journey', 'digital_gaps', 'digital_transformation', 'ai_use_cases']
    for section_key in standard_sections:
        if section_key in wave1_results:
            state.report_sections_ar[section_key] = {
                'heading': wave1_tasks[section_key][1],
                'raw_data': wave1_results[section_key],
            }

    print("[GenSections] ✓ Wave 1 results merged")

    # ──────────────────────────────────────────────────────────────────────────────────
    # WAVE 2: improvement_roadmap (depends on ai_use_cases from Wave 1)
    # ──────────────────────────────────────────────────────────────────────────────────
    print("[GenSections] WAVE 2: Generating Improvement Roadmap (depends on AI Use Cases)...")
    roadmap = generate_improvement_roadmap_section(state, api_key)
    state.report_sections_ar['improvement_roadmap'] = {
        'heading': 'ثامناً: خارطة الطريق التحسينية المقترحة',
        'raw_data': roadmap,
    }
    print("[GenSections] ✓ WAVE 2 complete")

    # ──────────────────────────────────────────────────────────────────────────────────
    # WAVE 3: conclusion (depends on both ai_use_cases and improvement_roadmap)
    # ──────────────────────────────────────────────────────────────────────────────────
    print("[GenSections] WAVE 3: Generating Conclusion (depends on AI Use Cases + Roadmap)...")
    conclusion = generate_conclusion_section(state, api_key)
    state.report_sections_ar['conclusion'] = {
        'heading': 'تاسعاً: الخلاصة — من البيانات إلى القرار',
        'raw_data': conclusion,
    }
    print("[GenSections] ✓ WAVE 3 complete")

    print("[Report Gen] ✅ All report sections generated successfully")


def _fix_unescaped_newlines(json_str: str) -> str:
    """Fix unescaped newlines in JSON string values (from LLM responses).

    When Claude returns JSON with literal newlines in Arabic text,
    this escapes them properly so json.loads() can parse it.
    """
    result = []
    i = 0
    in_string = False
    escape_next = False

    while i < len(json_str):
        char = json_str[i]

        if escape_next:
            result.append(char)
            escape_next = False
            i += 1
            continue

        if char == '\\' and in_string:
            result.append(char)
            escape_next = True
            i += 1
            continue

        if char == '"':
            in_string = not in_string
            result.append(char)
            i += 1
            continue

        if char == '\n' and in_string:
            # Replace literal newline with escaped version
            result.append('\\n')
            i += 1
            continue

        result.append(char)
        i += 1

    return ''.join(result)


def _build_pre_computed_findings(
    total_cases: int,
    misclassification_count: int,
    misclassification_rate: float,
    dominant_type: str,
    dominant_type_count: int,
    dominant_type_pct: float,
    complaint_subcategories: list,
    friction_points: list,
    friction_count: int,
    sla_closed: int,
    sla_rate: float,
) -> list:
    """
    Pre-compute the findings table deterministically from state data.

    This avoids the ambiguity where the LLM combined case counts from multiple friction points
    and readers couldn't distinguish between "N friction points" and "N cases".

    Each finding row is built from a template where:
    - The title (الاكتشاف) is ALWAYS deterministic from state, never LLM-generated
    - Case counts are separated from friction point counts in the description
    - Friction point groupings are explicit: "نقطتا احتكاك" (2 points) vs "N حالة" (N cases)

    Returns list of dicts with keys: number, title, description, importance
    """
    findings = []

    # ROW 1 — Classification accuracy gap
    findings.append({
        "number": 1,
        "title": f"تصنيف غير دقيق بنسبة {misclassification_rate:.1f}%",
        "description": (
            f"كانت {misclassification_count} من {total_cases} حالة مُصنَّفة أصلاً بشكل غير صحيح. "
            f"تمثل هذه الفجوة أساس التحديات التشغيلية المُكتشفة في هذا التحليل."
        ),
        "importance": "🔴 حرجة"
    })

    # ROW 2 — Dominant contact type dominance
    top_2_complaints = complaint_subcategories[:2]
    complaint_text = ""
    if len(top_2_complaints) >= 2:
        complaint_text = (
            f"أكبرها: {top_2_complaints[0]['name']} ({top_2_complaints[0]['count']} حالة) و"
            f"{top_2_complaints[1]['name']} ({top_2_complaints[1]['count']} حالة)"
        )
    elif len(top_2_complaints) == 1:
        complaint_text = f"{top_2_complaints[0]['name']} ({top_2_complaints[0]['count']} حالة)"

    findings.append({
        "number": 2,
        "title": f"الشكاوى تهيمن بـ {dominant_type_pct:.1f}% على عبء العمل",
        "description": (
            f"{dominant_type_count} حالة من {total_cases} كانت شكاوى بعد إعادة التصنيف. "
            f"{complaint_text}. هذا التركيز يعكس محور التحسين الأساسي للعمليات."
        ),
        "importance": "🔴 حرجة"
    })

    # ROW 3 — Largest friction point (SINGLE POINT, not grouped)
    if friction_points:
        top_friction = friction_points[0]
        findings.append({
            "number": 3,
            "title": f"نقطة احتكاك واحدة: {top_friction['name']} ({top_friction['case_count']} حالة، {round(top_friction['case_count']/total_cases*100, 1)}%)",
            "description": (
                f"تؤثر هذه النقطة على {top_friction['case_count']} حالة ({round(top_friction['case_count']/total_cases*100, 1)}% من الإجمالي). "
                f"السبب الجذري: {_root_cause_label(top_friction['root_cause'])}"
            ),
            "importance": top_friction['gap_severity']
        })

    # ROW 4 — Multiple friction points grouped by shared root cause
    # FIX: Use friction point COUNT, not case count, as the leading number
    if len(friction_points) >= 2:
        grouped_by_cause = {}
        for fp in friction_points[1:]:
            cause = fp['root_cause']
            if cause not in grouped_by_cause:
                grouped_by_cause[cause] = []
            grouped_by_cause[cause].append(fp)

        # Pick the largest group
        largest_group = max(grouped_by_cause.values(), key=lambda g: sum(f['case_count'] for f in g))
        friction_point_count_in_group = len(largest_group)
        case_count_in_group = sum(f['case_count'] for f in largest_group)
        group_cause = largest_group[0]['root_cause']
        group_cause_label = _root_cause_label(group_cause)

        # Title: Leading number is FRICTION POINT COUNT, not case count
        if friction_point_count_in_group == 1:
            point_label = "نقطة احتكاك واحدة"
        elif friction_point_count_in_group == 2:
            point_label = "نقطتا احتكاك"
        else:
            point_label = f"{friction_point_count_in_group} نقاط احتكاك"

        # Build point list for description
        point_names = ", ".join([f['name'] for f in largest_group])

        findings.append({
            "number": 4,
            "title": f"{point_label} — {group_cause_label} ({case_count_in_group} حالة)",
            "description": (
                f"تشترك {point_label} في سبب جذري مشترك: {point_names}. "
                f"إجمالي الحالات المتأثرة: {case_count_in_group} حالة. "
                f"تحسين هذه المجموعة سيسهم بشكل كبير في تقليل الاحتكاكات."
            ),
            "importance": largest_group[0]['gap_severity']
        })

    # ROW 5 — SLA / operational performance
    findings.append({
        "number": 5,
        "title": f"{sla_rate:.1f}% إغلاق في الوقت المحدد",
        "description": (
            f"تم إغلاق {sla_closed} من {total_cases} حالة في الوقت المحدد (معدل {sla_rate:.1f}%). "
            f"هذا يثبت القدرة التشغيلية، لكنه لا يعالج الفجوات الهيكلية في الدقة والفهم."
        ),
        "importance": "🟢 إيجابية"
    })

    return findings


def _root_cause_label(root_cause_category: str) -> str:
    """Map root cause category to Arabic label."""
    mapping = {
        'missing_info': 'غياب معلومات من الدليل',
        'inaccessible_info': 'معلومات موجودة لكنها صعبة الوصول',
        'no_proactive_notification': 'غياب الإشعار الاستباقي',
        'platform_bug': 'خلل تقني في المنصة',
        'policy_complexity': 'تعقيد إجراءات السياسة'
    }
    return mapping.get(root_cause_category, root_cause_category)


def generate_executive_summary_section(state: PipelineState, api_key: str) -> Dict[str, Any]:
    """
    Generate the executive summary section using Claude API.

    Extracts all necessary metrics from state and calls Claude with the detailed prompt.

    Args:
        state: Pipeline state with classified cases and analysis results
        api_key: Anthropic API key

    Returns:
        Dict with section_key and content following the specified JSON structure
    """
    try:
        # Extract classification metrics
        total_cases = state.total_cases
        all_classified = state.all_classified or []

        # Calculate distribution (after reclassification)
        # ISSUE 1 FIX: Ensure comparing final top_level against original case_type
        distribution = {}
        original_distribution = {}

        for case in all_classified:
            final_type = case.top_level  # Final classification (after Stage 2 + Stage 3)
            distribution[final_type] = distribution.get(final_type, 0) + 1

            # Track misclassifications against original CRM label
            original_type = case.case_type  # Original CRM نوع_المكالمة label
            original_distribution[original_type] = original_distribution.get(original_type, 0) + 1

        misclassification_count = state.reclassified_count
        misclassification_rate = state.reclassification_rate
        matched_original_count = total_cases - misclassification_count
        matched_original_rate = (matched_original_count / total_cases * 100) if total_cases > 0 else 0

        # Find dominant contact type
        dominant_type = max(distribution.items(), key=lambda x: x[1])[0] if distribution else ""
        dominant_type_count = distribution.get(dominant_type, 0)
        dominant_type_pct = (dominant_type_count / total_cases * 100) if total_cases > 0 else 0

        # Extract complaint subcategories from patterns
        complaint_subcategories = []
        if state.patterns:
            complaint_patterns = [p for p in state.patterns if 'شكوى' in (p.cluster_ar or p.cluster)]
            complaint_patterns.sort(key=lambda x: x.case_count, reverse=True)
            for pattern in complaint_patterns[:10]:
                pct = (pattern.case_count / dominant_type_count * 100) if dominant_type_count > 0 else 0
                complaint_subcategories.append({
                    'name': pattern.sub_theme_ar or pattern.sub_theme,
                    'count': pattern.case_count,
                    'pct_of_complaints': pct
                })

        # Extract friction points
        friction_points = []
        top_friction_point_name = ""
        top_friction_case_count = 0
        top_friction_pct = 0.0

        if state.journey_map:
            for friction in state.journey_map[:10]:
                # Map severity from root_cause_category
                severity_map = {
                    'missing_info': '🔴 حرجة',
                    'inaccessible_info': '🔴 حرجة',
                    'no_proactive_notification': '🟡 عالية',
                    'platform_bug': '🔴 حرجة',
                    'policy_complexity': '🟡 عالية'
                }
                severity = severity_map.get(friction.root_cause_category, '🟢 متوسطة')

                friction_points.append({
                    'name': friction.friction_point_ar or friction.friction_point,
                    'case_count': friction.case_count,
                    'root_cause': friction.root_cause_category,
                    'gap_severity': severity
                })

            # Extract top friction for direct injection into prompt
            if friction_points:
                top_friction_point_name = friction_points[0]['name']
                top_friction_case_count = friction_points[0]['case_count']
                top_friction_pct = round(top_friction_case_count / total_cases * 100, 1) if total_cases > 0 else 0.0

        # Calculate SLA metrics — check for 'نعم' (yes) in SLA compliance field
        # Must be done BEFORE _build_pre_computed_findings call
        sla_closed = sum(1 for c in all_classified if c.sla_color == 'نعم')
        sla_rate = (sla_closed / total_cases * 100) if total_cases > 0 else 0

        # FIX: Pre-compute findings table deterministically from state before LLM call
        # Avoids LLM inventing ambiguous case-count titles
        friction_count = len(state.journey_map or [])
        pre_computed_findings = _build_pre_computed_findings(
            total_cases=total_cases,
            misclassification_count=misclassification_count,
            misclassification_rate=misclassification_rate,
            dominant_type=dominant_type,
            dominant_type_count=dominant_type_count,
            dominant_type_pct=dominant_type_pct,
            complaint_subcategories=complaint_subcategories,
            friction_points=friction_points,
            friction_count=friction_count,
            sla_closed=sla_closed,
            sla_rate=sla_rate,
        )

        # Extract gaps with enhanced guidebook intelligence
        gap_table = []
        guidebook_coverage_metrics = {
            'total_gaps': 0,
            'covered': 0,
            'partially_covered': 0,
            'missing': 0,
            'avg_coverage_percentage': 0,
            'avg_match_confidence': 0,
            'gaps_with_proactive_notification_opportunity': 0
        }

        if state.gap_table:
            coverage_percentages = []
            match_confidences = []

            for gap in state.gap_table[:20]:
                severity_emoji = '🔴 حرجة' if gap.severity == 'Critical' else '🟡 عالية' if gap.severity == 'Medium' else '🟢 متوسطة'

                # Track coverage metrics
                guidebook_coverage_metrics['total_gaps'] += 1
                if gap.guidebook_status == 'Covered':
                    guidebook_coverage_metrics['covered'] += 1
                elif gap.guidebook_status == 'Partially Covered':
                    guidebook_coverage_metrics['partially_covered'] += 1
                else:
                    guidebook_coverage_metrics['missing'] += 1

                if gap.coverage_percentage:
                    coverage_percentages.append(gap.coverage_percentage)
                if gap.guidebook_match_confidence:
                    match_confidences.append(gap.guidebook_match_confidence)
                if gap.proactive_notification_opportunity:
                    guidebook_coverage_metrics['gaps_with_proactive_notification_opportunity'] += 1

                gap_table.append({
                    'topic': gap.topic_ar or gap.topic,
                    'case_count': gap.case_count,
                    'current_status': gap.guidebook_status,
                    'gap_type': gap.gap_type_ar or gap.gap_type,
                    'coverage_percentage': gap.coverage_percentage,
                    'clarity': gap.clarity_assessment,
                    'format': gap.format_assessment,
                    'has_visuals': gap.has_visual_guidance,
                    'match_confidence': gap.guidebook_match_confidence,
                    'can_use_proactive_notification': gap.proactive_notification_opportunity,
                    'guidebook_excerpt': gap.guidebook_excerpt_ar or gap.guidebook_excerpt,
                    'recommendation': gap.recommendation
                })

            # Calculate averages
            if coverage_percentages:
                guidebook_coverage_metrics['avg_coverage_percentage'] = sum(coverage_percentages) / len(coverage_percentages)
            if match_confidences:
                guidebook_coverage_metrics['avg_match_confidence'] = sum(match_confidences) / len(match_confidences)

        # Calculate proactive notification impact (from notification opportunities)
        proactive_notification_total = 0
        if state.notification_opportunities:
            proactive_notification_total = sum(n.get('cases_eliminated', n.get('case_count', 0)) for n in state.notification_opportunities)
        proactive_notification_pct = (proactive_notification_total / total_cases * 100) if total_cases > 0 else 0

        # Digital channel percentage
        digital_cases = sum(1 for c in all_classified if c.case_channel in ['app', 'web', 'website', 'application'])
        digital_channel_pct = (digital_cases / total_cases * 100) if total_cases > 0 else 0

        # Date range extraction (placeholder - would be parsed from data in real scenario)
        date_range = convert_month_year_to_arabic(state.month_year) or "يناير — مارس 2026"
        quarter_label = "Q1 2026"  # Would be calculated from month_year

        # Serialize pre-computed findings for the prompt
        findings_json = json.dumps(pre_computed_findings, ensure_ascii=False, indent=2)

        # Build the prompt with PRE-COMPUTED findings table (locked, not LLM-generated)
        prompt = f"""You are an expert CX strategist writing the executive summary of a formal Arabic government report on customer inquiry analysis.

OUTPUT LANGUAGE: Arabic only. Do not generate English translations.

INPUTS PROVIDED:
- total_cases: {total_cases}
- date_range: {date_range}
- friction_count: {friction_count} (total distinct friction points identified)
- gap_table: {json.dumps(gap_table, ensure_ascii=False)}
- guidebook_coverage_metrics: {json.dumps(guidebook_coverage_metrics, ensure_ascii=False)}
- sla_rate: {sla_rate:.1f}%

KEY STRUCTURAL INSIGHT FOR YOUR REFERENCE:
The main reclassification finding: {misclassification_rate:.1f}% of cases were initially misclassified.
When corrected, {dominant_type} rises to {dominant_type_pct:.1f}% of total workload ({dominant_type_count} cases).

YOUR TASK — WRITE ONLY TWO SECTIONS:
─────────────────────────────────────────────

1. FRAMING PARAGRAPH (فقرة الإطار)
   Write 2–3 sentences that:
   - State how many cases were analyzed, from which CRM source, and for which period
   - Name both data sources: "تصدير نظام إدارة علاقات العملاء" and "دليل الخدمات والأسئلة الشائعة"
   - Declare the report's purpose: transforming data into actionable decisions, not just presenting numbers
   - End with the most striking structural discovery: the {dominant_type} reclassification finding

   Style: Open with "يُقدّم هذا التقرير...". Third sentence must begin with
   "المُستجد الجوهري:" and deliver insight, not description.

2. الرسالة الجوهرية (CORE MESSAGE)
   One paragraph placed AFTER the locked findings table below. Must:
   - Begin with "الرسالة الجوهرية:"
   - Name the specific systemic failure: {misclassification_rate:.1f}% misclassification + {friction_count} distinct friction points
   - State what fixing it unlocks, using specific numbers from the inputs
   - End with a forward-looking statement about data-driven strategy

   Style: 3 sentences maximum. No bullet points. Assertive, not descriptive.

─────────────────────────────────────────────
LOCKED FINDINGS TABLE — COPY VERBATIM, DO NOT MODIFY:
─────────────────────────────────────────────

This table is LOCKED and pre-computed from authoritative state data.
Your ONLY task is to write framing_paragraph and core_message.
Do NOT invent, modify, or re-order the findings table rows.
Do NOT change any number in the table — case counts, percentages, or friction point descriptions.

Pre-computed key_findings table (5 rows):
{findings_json}

CRITICAL DISTINCTION (重要):
- Friction point COUNT = total number of distinct access barriers (نقاط احتكاك)
- Case COUNT = number of cases affected by that friction point (حالات)
- These are different numbers. Do NOT confuse them.
- Example: If ROW 4 says "نقطتا احتكاك — ... (10 حالة)", that means:
  - 2 friction points (نقطتا احتكاك)
  - Affecting 10 cases total (10 حالة)
  - NOT 10 friction points

VALIDATION RULE:
If any finding headline number exceeds {friction_count}, that finding is invalid.
Example: "10 نقاط احتكاك" is invalid if friction_count={friction_count}.
(This validates that we're counting the right dimension.)

─────────────────────────────────────────────
OUTPUT FORMAT — return JSON with THREE fields:
─────────────────────────────────────────────
{{
  "section": "executive_summary",
  "framing_paragraph": "...",  ← Your writing (2–3 sentences, Arabic only)
  "key_findings": {findings_json},  ← LOCKED: return as provided above, no changes
  "core_message": "..."  ← Your writing (1 paragraph, 3 sentences max)
}}

RULES:
- key_findings: Return EXACTLY as provided above — every row, every value, unchanged.
- framing_paragraph: Arabic only, 2–3 sentences, cite date_range and both data sources.
- core_message: Arabic only, 3 sentences max, begin with "الرسالة الجوهرية:".
- No markdown, no extra keys, no extra nesting.
- CRITICAL: Do NOT use double-quote characters (") inside any string value. Use « » for citations.
"""

        client = anthropic.Anthropic(api_key=api_key)
        print(f"[ExecSummary] Calling API with model claude-sonnet-4-6")
        print(f"[ExecSummary] total_cases={total_cases}, reclassified={misclassification_count}")
        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4000,  # Increased to handle full response
            messages=[
                {
                    "role": "user",
                    "content": prompt
                }
            ]
        )

        # Extract JSON from response
        response_text = message.content[0].text

        # Try to parse JSON from response
        import re

        # First try: look for ```json ... ``` block (with or without closing ```)
        # Handle both complete and incomplete code blocks
        result = None
        json_code_block = re.search(r'```\s*(?:json)?\s*\n(.*?)(?:\n```|$)', response_text, re.DOTALL)
        if json_code_block:
            json_candidate = json_code_block.group(1).strip()
            try:
                result = json.loads(json_candidate)
            except json.JSONDecodeError:
                pass  # Fall through to next method

        # Second try: extract between first { and last } using state machine
        if not result:
            first_brace = response_text.find('{')
            if first_brace != -1:
                # Use state machine to find matching closing brace
                depth = 0
                in_string = False
                escape = False

                for i in range(first_brace, len(response_text)):
                    char = response_text[i]

                    # Handle escape sequences
                    if escape:
                        escape = False
                        continue

                    if char == '\\' and in_string:
                        escape = True
                        continue

                    # Track string boundaries (only outside strings do braces matter)
                    if char == '"':
                        in_string = not in_string
                        continue

                    if in_string:
                        continue

                    # Track brace depth
                    if char == '{':
                        depth += 1
                    elif char == '}':
                        depth -= 1
                        if depth == 0:
                            # Found closing brace
                            json_str = response_text[first_brace:i + 1]
                            try:
                                result = json.loads(json_str)
                                break
                            except json.JSONDecodeError as e:
                                # Try to fix unescaped newlines in JSON strings
                                fixed_json = _fix_unescaped_newlines(json_str)
                                try:
                                    result = json.loads(fixed_json)
                                    break
                                except json.JSONDecodeError:
                                    raise RuntimeError(
                                        f"[ExecSummary] Failed to parse executive summary JSON: {e}\n"
                                        f"Attempted string length: {len(json_str)}\n"
                                        f"First 500 chars: {json_str[:500]}\n"
                                        f"Last 500 chars: {json_str[-500:] if len(json_str) > 500 else 'N/A'}"
                                    )

        if not result:
            print("No JSON found in executive summary response")
            print(f"Response first 500 chars: {response_text[:500]}")
            raise RuntimeError("Executive summary: No JSON found in API response")

        # ── FIX: Validate and reinject pre-computed findings ────────────────────────
        # The LLM should NOT have modified the findings table, but we validate and reinject anyway
        if 'key_findings' in result and isinstance(result['key_findings'], list):
            if len(result['key_findings']) != len(pre_computed_findings):
                print(
                    f"[ExecSummary] WARNING: LLM returned {len(result['key_findings'])} findings, "
                    f"expected {len(pre_computed_findings)}. Reinjecting pre-computed findings."
                )
            # Validate that leading numbers don't exceed friction_count
            for finding in result.get('key_findings', []):
                title = finding.get('title', '')
                # Extract first number from title (simple heuristic)
                import re as re_module
                numbers = re_module.findall(r'\d+', title)
                if numbers:
                    first_num = int(numbers[0])
                    if first_num > friction_count:
                        print(
                            f"[ExecSummary] WARNING: Finding title contains number {first_num} "
                            f"which exceeds friction_count={friction_count}. "
                            f"This suggests confusion between friction points and case counts. "
                            f"Reinjecting pre-computed findings to ensure consistency."
                        )
                        break

        # Reinject pre-computed findings
        result['key_findings'] = pre_computed_findings

        print(
            f"[ExecSummary] ✅ OK — "
            f"framing_paragraph_len={len(result.get('framing_paragraph', ''))}, "
            f"key_findings_rows={len(result.get('key_findings', []))}, "
            f"core_message_len={len(result.get('core_message', ''))}"
        )
        return result

    except Exception as e:
        print(f"[ExecSummary] ❌ Error: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        raise  # Don't silently return None — let caller see the error


def _is_valid_table(t: dict) -> bool:
    """Validate that a table dict has required structure with non-empty rows."""
    return (isinstance(t, dict)
            and isinstance(t.get('rows'), list)
            and len(t.get('rows', [])) > 0
            and len(t.get('columns', [])) > 0)


def _build_sources_table(state: PipelineState, sources_input: List[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Build bilingual sources table for methodology section.

    BUG 1 FIX: If the LLM returns a plain Arabic-only list (sources_input),
    use it as rows_ar and build corresponding rows_en from state values.
    """
    columns_ar = ["المصدر", "الطبيعة", "الحجم", "الفترة"]
    columns_en = ["Source", "Nature", "Size", "Period"]

    if not state or not state.total_cases:
        raise ValueError("state.total_cases not set - Stage 1 validation must complete successfully")
    if not state.month_year:
        raise ValueError("state.month_year not set - Stage 1 validation must complete successfully")
    if not hasattr(state, 'guidebook_pages') or state.guidebook_pages is None:
        raise ValueError("state.guidebook_pages not set - Stage 5 must complete successfully")
    if not hasattr(state, 'guidebook_faq_count') or state.guidebook_faq_count is None:
        raise ValueError("state.guidebook_faq_count not set - Stage 5 must complete successfully")
    if not hasattr(state, 'guidebook_year') or state.guidebook_year is None:
        raise ValueError("state.guidebook_year not set - Stage 5 must complete successfully")

    # Use closed_cases_count (where تاريخ_إغلاق_الطلب is not empty)
    case_count = state.closed_cases_count if state.closed_cases_count > 0 else state.total_cases
    date_range = convert_month_year_to_arabic(state.month_year)
    guidebook_pages = state.guidebook_pages
    guidebook_faq_count = state.guidebook_faq_count
    guidebook_year = state.guidebook_year

    # BUG 1: Check if LLM returned a plain Arabic-only list
    if (isinstance(sources_input, list) and len(sources_input) >= 2
            and isinstance(sources_input[0], dict)
            and 'المصدر' in sources_input[0]):
        # Use LLM's Arabic rows and build English equivalents from state
        rows_ar = sources_input[:2]
        rows_en = [
            {
                'Source': 'Inquiry Analysis',
                'Nature': 'CRM data with unstructured text fields (case details, resolutions, service names, case descriptions)',
                'Size': f'{case_count} closed cases',
                'Period': date_range
            },
            {
                'Source': 'Customer Services Guidebook',
                'Nature': 'Official reference covering traffic, licensing, and security services, plus FAQ validation and gap analysis',
                'Size': f'{guidebook_pages} pages, {guidebook_faq_count} FAQs',
                'Period': guidebook_year
            }
        ]
    else:
        # Fallback: build default structure
        rows_ar = [
            {
                "المصدر": "تحليل الاستفسارات",
                "الطبيعة": "بيانات CRM — نصوص غير مهيكلة (تفاصيل الطلب، الحلول، أسماء الخدمات، أوصاف الحالات)",
                "الحجم": f"{case_count} حالة مغلقة",
                "الفترة": date_range
            },
            {
                "المصدر": "دليل خدمات العملاء",
                "الطبيعة": "الدليل الرسمي يغطي خدمات المرور والترخيص والأمن، مع التحقق من الأسئلة الشائعة وتحليل فجوات التغطية",
                "الحجم": f"{guidebook_pages} صفحة، {guidebook_faq_count} سؤالاً",
                "الفترة": guidebook_year
            }
        ]

        rows_en = [
            {
                "Source": "Inquiry Analysis",
                "Nature": "CRM data with unstructured text fields (case details, resolutions, service names, case descriptions)",
                "Size": f"{case_count} closed cases",
                "Period": date_range
            },
            {
                "Source": "Customer Services Guidebook",
                "Nature": "Official reference covering traffic, licensing, and security services, plus FAQ validation and gap analysis",
                "Size": f"{guidebook_pages} pages, {guidebook_faq_count} FAQs",
                "Period": guidebook_year
            }
        ]

    return {
        'columns_ar': columns_ar,
        'columns_en': columns_en,
        'rows_ar': rows_ar,
        'rows_en': rows_en
    }


def build_bilingual_report_sections(exec_summary: Dict[str, Any], methodology: Dict[str, Any], state: PipelineState = None) -> tuple[Dict[str, Any], Dict[str, Any]]:
    """
    Transform flat LLM responses into nested bilingual section structure.

    Returns two dicts:
    - report_sections_ar: all content in Arabic
    - report_sections_en: all content in English

    Each dict has structure: {"sections": [...]}
    with nested subsections matching the expected format.
    """
    report_sections_ar = {"sections": []}
    report_sections_en = {"sections": []}

    # ========== SECTION 1: Executive Summary ==========
    if exec_summary:
        # Build key findings table for AR version
        key_findings_table_ar = {
            "columns": ["#", "الاكتشاف", "الوصف", "مستوى الأهمية"],
            "rows": [],
            "row_count": 0,
            "col_count": 4,
            "original_index": 0
        }

        # Build key findings table for EN version
        key_findings_table_en = {
            "columns": ["#", "Discovery", "Description", "Importance Level"],
            "rows": [],
            "row_count": 0,
            "col_count": 4,
            "original_index": 0
        }

        # Convert key_findings array to table rows
        if "key_findings" in exec_summary and isinstance(exec_summary["key_findings"], list):
            for finding in exec_summary["key_findings"]:
                number = finding.get("number", "")

                # AR row
                key_findings_table_ar["rows"].append({
                    "#": str(number),
                    "الاكتشاف": finding.get("title_ar", ""),
                    "الوصف": finding.get("description_ar", ""),
                    "مستوى الأهمية": finding.get("importance_ar", "")
                })

                # EN row
                key_findings_table_en["rows"].append({
                    "#": str(number),
                    "Discovery": finding.get("title_en", ""),
                    "Description": finding.get("description_en", ""),
                    "Importance Level": finding.get("importance_en", "")
                })

        key_findings_table_ar["row_count"] = len(key_findings_table_ar["rows"])
        key_findings_table_en["row_count"] = len(key_findings_table_en["rows"])

        # Executive Summary section - Arabic version
        exec_summary_ar = {
            "id": "section_15_أولا_الملخص_التنفيذي",
            "title": "أولاً: الملخص التنفيذي — التحليلات الرئيسية",
            "title_en": "Executive Summary",
            "level": 2,
            "content": exec_summary.get("framing_paragraph_ar", ""),
            "tables": [],
            "charts": [],
            "subsections": [
                {
                    "id": "section_16_النتائج_الرئيسية",
                    "title": "النتائج الرئيسية",
                    "title_en": "Key Findings",
                    "level": 3,
                    "content": exec_summary.get("core_message_ar", ""),
                    "tables": [key_findings_table_ar] if key_findings_table_ar["rows"] else [],
                    "charts": []
                }
            ]
        }

        # Executive Summary section - English version
        exec_summary_en = {
            "id": "section_15_أولا_الملخص_التنفيذي",
            "title": "Executive Summary",
            "title_en": "Executive Summary",
            "level": 2,
            "content": exec_summary.get("framing_paragraph_en", ""),
            "tables": [],
            "charts": [],
            "subsections": [
                {
                    "id": "section_16_النتائج_الرئيسية",
                    "title": "Key Findings",
                    "title_en": "Key Findings",
                    "level": 3,
                    "content": exec_summary.get("core_message_en", ""),
                    "tables": [key_findings_table_en] if key_findings_table_en["rows"] else [],
                    "charts": []
                }
            ]
        }

        report_sections_ar["sections"].append(exec_summary_ar)
        report_sections_en["sections"].append(exec_summary_en)

    # ========== SECTION 2: Methodology ==========
    if methodology:
        # Build sources table from LLM response
        sources_raw = _build_sources_table(state, methodology.get("sources_table_ar", []))

        # Methodology section - Arabic version
        sources_table_ar = {
            "columns": sources_raw["columns_ar"],
            "rows": sources_raw["rows_ar"],
            "row_count": len(sources_raw["rows_ar"]),
            "col_count": len(sources_raw["columns_ar"]),
            "original_index": 0
        }

        sources_table_en = {
            "columns": sources_raw["columns_en"],
            "rows": sources_raw["rows_en"],
            "row_count": len(sources_raw["rows_en"]),
            "col_count": len(sources_raw["columns_en"]),
            "original_index": 0
        }

        methodology_ar = {
            "id": "section_17_ثانيا_المنهجية_وطبيعة",
            "title": "ثانياً: المنهجية وطبيعة المصادر",
            "title_en": "Methodology and Data Sources",
            "level": 2,
            "content": "",
            "tables": [],
            "charts": [],
            "subsections": [
                {
                    "id": "section_18_21_المصادر_المحللة",
                    "title": "2.1 المصادر المُحلَّلة",
                    "title_en": "2.1 Sources Analyzed",
                    "level": 3,
                    "content": "",
                    "tables": [sources_table_ar] if _is_valid_table(sources_table_ar) else [],
                    "charts": []
                },
                {
                    "id": "section_19_22_منهجية_التصنيف",
                    "title": "2.2 منهجية التصنيف",
                    "title_en": "2.2 Classification Methodology",
                    "level": 3,
                    "content": methodology.get("classification_method_ar", ""),
                    "tables": [],
                    "charts": []
                },
                {
                    "id": "section_20_23_الحقول_المحللة",
                    "title": "2.3 الحقول المُحلَّلة",
                    "title_en": "2.3 Analyzed Fields",
                    "level": 3,
                    "content": methodology.get("analyzed_fields_ar", ""),
                    "tables": [],
                    "charts": []
                }
            ]
        }

        methodology_en = {
            "id": "section_17_ثانيا_المنهجية_وطبيعة",
            "title": "Methodology and Data Sources",
            "title_en": "Methodology and Data Sources",
            "level": 2,
            "content": "",
            "tables": [],
            "charts": [],
            "subsections": [
                {
                    "id": "section_18_21_المصادر_المحللة",
                    "title": "2.1 Sources Analyzed",
                    "title_en": "2.1 Sources Analyzed",
                    "level": 3,
                    "content": "",
                    "tables": [sources_table_en] if _is_valid_table(sources_table_en) else [],
                    "charts": []
                },
                {
                    "id": "section_19_22_منهجية_التصنيف",
                    "title": "2.2 Classification Methodology",
                    "title_en": "2.2 Classification Methodology",
                    "level": 3,
                    "content": methodology.get("classification_method_en", ""),
                    "tables": [],
                    "charts": []
                },
                {
                    "id": "section_20_23_الحقول_المحللة",
                    "title": "2.3 Analyzed Fields",
                    "title_en": "2.3 Analyzed Fields",
                    "level": 3,
                    "content": methodology.get("analyzed_fields_en", ""),
                    "tables": [],
                    "charts": []
                }
            ]
        }

        report_sections_ar["sections"].append(methodology_ar)
        report_sections_en["sections"].append(methodology_en)

    return report_sections_ar, report_sections_en


def generate_methodology_section(state: PipelineState, api_key: str) -> Dict[str, Any]:
    """
    Generate the methodology section using Claude API.

    Covers:
    - 2.1 المصادر المُحلَّلة (Sources analyzed table)
    - 2.2 منهجية التصنيف (Classification methodology prose)
    - 2.3 الحقول المُحلَّلة (Analyzed fields prose)

    Args:
        state: Pipeline state with classified cases and analysis results
        api_key: Anthropic API key

    Returns:
        Dict with section_key and content following the specified JSON structure
    """
    try:
        # Extract classification metrics
        total_cases = state.total_cases
        # Use closed_cases_count (where تاريخ_إغلاق_الطلب is not empty) for "حالة مغلقة" reporting
        closed_cases = state.closed_cases_count if state.closed_cases_count > 0 else total_cases
        all_classified = state.all_classified or []

        # Use centralized reclassification stats from state (computed in generate_artifacts_stage6)
        misclassification_count = state.reclassified_count
        misclassification_rate = state.reclassification_rate
        matched_original_count = closed_cases - misclassification_count
        matched_original_rate = (matched_original_count / closed_cases * 100) if closed_cases > 0 else 0

        # Guidebook stats (must be set by Stage 5)
        if not hasattr(state, 'guidebook_pages') or state.guidebook_pages is None:
            raise ValueError("guidebook_pages not set - Stage 5 must complete successfully")
        if not hasattr(state, 'guidebook_faq_count') or state.guidebook_faq_count is None:
            raise ValueError("guidebook_faq_count not set - Stage 5 must complete successfully")
        if not hasattr(state, 'guidebook_year') or state.guidebook_year is None:
            raise ValueError("guidebook_year not set - Stage 5 must complete successfully")

        guidebook_pages = state.guidebook_pages
        guidebook_faq_count = state.guidebook_faq_count
        guidebook_year = state.guidebook_year

        # Date range (must be set by Stage 1)
        if not state.month_year:
            raise ValueError("month_year not set - Stage 1 validation must complete successfully")
        date_range = convert_month_year_to_arabic(state.month_year)

        # Build two-level taxonomy structure
        # Top level types (4 required)
        top_level_types = ['شكوى', 'طلب', 'استفسار', 'شكر وثناء']

        # Load SUB_CLASSIFICATIONS from stage2_rules (required for taxonomy)
        try:
            from .stage2_rules import SUB_CLASSIFICATIONS
        except ImportError as e:
            raise ImportError(f"Cannot load SUB_CLASSIFICATIONS from stage2_rules: {e}")

        complaint_subs = SUB_CLASSIFICATIONS.get('شكوى', [])
        request_subs = SUB_CLASSIFICATIONS.get('طلب', [])
        inquiry_subs = SUB_CLASSIFICATIONS.get('استفسار', [])
        praise_sub = SUB_CLASSIFICATIONS.get('شكر وثناء', [''])[-1] if 'شكر وثناء' in SUB_CLASSIFICATIONS else 'شكر وثناء'

        if not complaint_subs:
            raise ValueError("SUB_CLASSIFICATIONS['شكوى'] is empty - stage2_rules configuration is invalid")
        if not request_subs:
            raise ValueError("SUB_CLASSIFICATIONS['طلب'] is empty - stage2_rules configuration is invalid")
        if not inquiry_subs:
            raise ValueError("SUB_CLASSIFICATIONS['استفسار'] is empty - stage2_rules configuration is invalid")

        # Classification logic priority tree (from stage2_rules)
        priority_tree = [
            "اعتراض على مخالفة مرورية",
            "تقديم بلاغ أمني أو مروري",
            "طلب تصريح سلاح أو ترخيص",
            "شكوى عن عدم استلام الخدمة",
            "شكوى على خطأ تقني أو في النظام",
            "شكوى على تأخر المعالجة",
            "استفسار عن الرخص والمركبات",
            "شكر وثناء",
        ]

        # Confidence thresholds
        confidence_threshold_stage2 = 0.75
        llm_confidence_threshold = 0.65

        # Use preserved queue counts from state (populated after stages 2 & 3)
        llm_queue_count = state.llm_queue_count
        human_review_count = state.human_review_count

        # Analyze unstructured fields
        desc_lengths = [len(c.case_title or '') for c in all_classified if c.case_title]
        res_lengths = [len(c.resolution_response or '') for c in all_classified if c.resolution_response]

        desc_avg_chars = int(sum(desc_lengths) / len(desc_lengths)) if desc_lengths else 250
        res_avg_chars = int(sum(res_lengths) / len(res_lengths)) if res_lengths else 300

        # Compute language distribution from case titles
        def detect_arabic_proportion(text: str) -> float:
            """Return proportion of Arabic characters in first 50 characters."""
            if not text:
                return 0.0
            sample = text[:50]
            arabic_count = sum(1 for ch in sample if '؀' <= ch <= 'ۿ')
            return arabic_count / len(sample) if sample else 0.0

        arabic_title_count = sum(1 for c in all_classified if detect_arabic_proportion(c.case_title) > 0.3)
        arabic_response_count = sum(1 for c in all_classified if detect_arabic_proportion(c.resolution_response) > 0.3)

        title_ar_pct = int(arabic_title_count / len(all_classified) * 100) if all_classified else 72
        response_ar_pct = int(arabic_response_count / len(all_classified) * 100) if all_classified else 72

        desc_lang_dist = f"{title_ar_pct}% عربي / {100 - title_ar_pct}% إنجليزي"
        res_lang_dist = f"{response_ar_pct}% عربي / {100 - response_ar_pct}% إنجليزي"

        # Calculate total sub-classification count
        total_sub_count = len(complaint_subs) + len(request_subs) + len(inquiry_subs) + 1

        # Use actual guidebook service categories (extracted from guidebook sections)
        if not state.guidebook_topics:
            raise ValueError("guidebook_topics not set - Stage 5 must complete successfully first")
        guidebook_topics_str = ', '.join(state.guidebook_topics[:5])  # Limit to 5 topics
        if len(state.guidebook_topics) > 5:
            guidebook_topics_str += f", and {len(state.guidebook_topics) - 5} more"

        # Compute pattern distribution by top_level type
        patterns_by_type = {}
        for pattern in state.patterns:
            top_level = pattern.top_level or pattern.cluster
            if top_level not in patterns_by_type:
                patterns_by_type[top_level] = 0
            patterns_by_type[top_level] += pattern.case_count

        patterns_dist = '\n    '.join(
            f"{tl}: {count} cases ({count/total_cases*100:.1f}%)"
            for tl, count in sorted(patterns_by_type.items(), key=lambda x: x[1], reverse=True)
        ) if patterns_by_type else "No patterns identified"

        # Build the prompt
        prompt = f"""You are documenting the methodology section of a formal Arabic government report for
Fujairah Police customer inquiry analysis.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
INPUTS PROVIDED TO YOU
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

source_1_crm:
  - closed_cases: {closed_cases}
  - date_range: "{date_range}"
  - content_description: >
      CRM data — unstructured text fields (case details, resolutions,
      service names, case descriptions). Only cases where تاريخ_إغلاق_الطلب is not empty.

source_2_guidebook:
  - pages: {guidebook_pages}
  - faq_count: {guidebook_faq_count}
  - validated_faq_candidates: {state.validated_faqs_count}
  - edition_year: "{guidebook_year}"
  - content_description: >
      Official customer services guidebook covering {guidebook_topics_str},
      plus FAQ validation and service gap analysis.

classification_stats:
  - closed_cases: {closed_cases}
  - reclassified_count: {misclassification_count}
  - reclassification_rate: "{misclassification_rate:.1f}%"
  - matched_original_count: {matched_original_count}
  - matched_original_rate: "{matched_original_rate:.1f}%"

two_level_taxonomy:
  top_level_types:
    - شكوى
    - طلب
    - استفسار
    - شكر وثناء
  sub_classifications:
    شكوى: {json.dumps(complaint_subs, ensure_ascii=False)}
    طلب: {json.dumps(request_subs, ensure_ascii=False)}
    استفسار: {json.dumps(inquiry_subs, ensure_ascii=False)}
    شكر وثناء: ["{praise_sub}"]

classification_logic:
  priority_tree:
    1: "اعتراض على مخالفة مرورية"
    2: "تقديم بلاغ أمني أو مروري"
    3: "طلب تصريح سلاح أو ترخيص"
    4: "شكوى عن عدم استلام الخدمة"
    5: "شكوى على خطأ تقني أو في النظام"
    6: "شكوى على تأخر المعالجة"
    7: "استفسار عن الرخص والمركبات"
    8: "شكر وثناء"
    default: "map from CRM نوع_المكالمة label if available"
  confidence_threshold_stage2: {confidence_threshold_stage2}
  fallthrough_confidence: 0.70–0.75

llm_stage:
  model: "claude-haiku-4-5-20251001"
  trigger: "confidence < {confidence_threshold_stage2}"
  llm_confidence_threshold: {llm_confidence_threshold}
  low_confidence_action: "routed to human review queue — excluded from report counts"
  cases_to_llm: {llm_queue_count}
  cases_to_human_review: {human_review_count}

pipeline_output_distribution:
  top_level_categories:
    {patterns_dist}

analyzed_fields:
  structured:
    - رقم_الطلب
    - تاريخ_الإنشاء
    - قناة_تقديم_الخدمة
    - الخدمة_الرئيسية
    - نوع_المكالمة
    - الجنسية
    - الإدارة المختصة
  unstructured:
    - field: تفاصيل_الطلب
      avg_chars: {desc_avg_chars}
      language_dist: "{desc_lang_dist}"
    - field: الحل
      avg_chars: {res_avg_chars}
      language_dist: "{res_lang_dist}"

core_definitional_principle: >
  طبيعة المطلوب — وليس الصياغة — هي المعيار الفاصل.
  الحالة التي تُعبّر عن استياء تُصنَّف شكوى حتى لو تضمّنت طلب إجراء.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
YOUR TASK
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Write the two subsections below in formal Arabic only. Do not generate English.
Do not invent numbers or details not present in the inputs above.

NOTE: Section 2.2 (منهجية التصنيف) is hardcoded and will be injected separately.
Only generate sections 2.1 and 2.3.

───────────────────────────────────────────────
2.1  المصادر المُحلَّلة
───────────────────────────────────────────────
Produce a 2-row Arabic table with columns:
  المصدر | الطبيعة | الحجم | الفترة

Row 1 — CRM inquiry data:
  - المصدر: تحليل الاستفسارات
  - الطبيعة: describe it as CRM data with unstructured text fields
              (case details, resolutions, service names, case descriptions)
  - الحجم: {closed_cases} حالة مغلقة
  - الفترة: {date_range}

Row 2 — Customer services guidebook:
  - المصدر: دليل خدمات العملاء
  - الطبيعة: describe its role covering {guidebook_topics_str},
              plus FAQ validation and gap analysis
  - الحجم: {guidebook_pages} صفحة، {state.validated_faqs_count} أسئلة مصدقة
  - الفترة: {guidebook_year}

No prose paragraphs in this subsection — the table IS the content.

───────────────────────────────────────────────
2.3  الحقول المُحلَّلة
───────────────────────────────────────────────
Write a prose paragraph in Arabic with two groups:

Group A — Structured fields (handled by dashboard/PowerBI, not this pipeline):
  List all fields from analyzed_fields.structured using their Arabic names.

Group B — Unstructured fields — the focus of this analysis:
  - تفاصيل_الطلب: avg {desc_avg_chars} chars, {desc_lang_dist}
  - الحل: avg {res_avg_chars} chars, {res_lang_dist}
  Close with one sentence explaining that these two fields reveal the
  true nature of each case beyond the CRM's original label (نوع_المكالمة).

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
OUTPUT FORMAT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Return a single flat JSON object with these exact keys:

{{
  "sources_table": [
    {{
      "المصدر": "...",
      "الطبيعة": "...",
      "الحجم": "...",
      "الفترة": "..."
    }},
    {{
      "المصدر": "...",
      "الطبيعة": "...",
      "الحجم": "...",
      "الفترة": "..."
    }}
  ],
  "analyzed_fields": "..."
}}

Rules:
- sources_table: exactly 2 row objects, Arabic column keys only
- analyzed_fields: single Arabic prose paragraph covering structured fields (list) then unstructured fields (تفاصيل_الطلب and الحل with avg chars and language distribution)
- No markdown, no extra keys, no nesting beyond what is shown above
- All content must be in Arabic only
- CRITICAL: Do NOT use double-quote characters (") inside any string value. Use « » for citations.
- Section 2.2 (classification methodology) is omitted — it will be hardcoded separately"""

        # HARDCODED SECTION 2.2: منهجية التصنيف
        classification_method_hardcoded = (
            "يعتمد التحليل على شجرة قرار من أربعة مستويات، حيث طبيعة المطلوب وليس الصياغة هي المعيار الفاصل. يُطرح على كل حالة اختباران متتاليان: "
            "(1) الاختبار الأول: هل يُعبّر النص عن استياء، أو إبلاغ عن إخفاق، أو رغبة في تقديم بلاغ رسمي أو اعتراض؟ "
            "(2) الاختبار الثاني: هل يطلب النص تنفيذ إجراء محدد كتقديم خدمة أو متابعة طلب أو تعديل بيانات، وليس مجرد الحصول على معلومة؟\n"
            "قواعد التصنيف:\n"
            "→ إن كان الجواب نعم على الأول: شكوى (حتى لو تضمّن النص طلباً لاتخاذ إجراء)؛\n"
            "→ إن كان الجواب نعم على الثاني فقط: طلب (دون أن يكون الغرض إبلاغاً عن مشكلة)؛\n"
            "→ إن كان الغرض سؤالاً أو استعلاماً عن معلومة: استفسار؛\n"
            "→ إن كان التعبير عن رضا وثناء: شكر وثناء."
        )

        # Build sources table using closed_cases_count (where تاريخ_إغلاق_الطلب is not empty)
        closed_count = state.closed_cases_count if state.closed_cases_count > 0 else total_cases
        sources_rows = [
            {
                "المصدر": "تحليل الاستفسارات",
                "الطبيعة": "بيانات CRM — نصوص غير مهيكلة (تفاصيل الحالة، الحلول، أسماء الخدمات، وصف الحالة)",
                "الحجم": f"{closed_count} حالة مغلقة",
                "الفترة": date_range
            },
            {
                "المصدر": "دليل خدمات العملاء",
                "الطبيعة": f"يغطي {guidebook_topics_str}، مع التحقق من الأسئلة الشائعة وتحليل الفجوات",
                "الحجم": f"{guidebook_pages} صفحة، {state.validated_faqs_count} أسئلة مصدقة",
                "الفترة": guidebook_year
            }
        ]

        # HARDCODED SECTION 2.3: الحقول المحللة
        analyzed_fields_text = (
            "الحقول المنظمة المستخدمة في التحليل: رقم الطلب، الخدمة الرئيسية، نوع المكالمة الأصلي، حالة الطلب، قناة التواصل، الجنسية، والإدارة المختصة. "
            "الحقول غير المنظمة — محور هذا التحليل: تفاصيل الطلب ورد المعالجة. هذان الحقلان يكشفان الطبيعة الحقيقية لكل حالة بعيداً عن التصنيف الشكلي في النظام."
        )

        return {
            "sources_table": sources_rows,
            "classification_method": classification_method_hardcoded,
            "analyzed_fields": analyzed_fields_text
        }

    except Exception as e:
        print(f"[Methodology] ❌ Error: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        raise  # Don't silently return None — let caller see the error


def _build_summary_context(state: PipelineState) -> str:
    """Build summary context for report generation."""
    lines = [
        f"Total Cases: {state.total_cases}",
        f"Date: {state.month_year or 'Q1 2026'}",
    ]

    # Add distribution
    if state.all_classified:
        type_counts = {}
        for case in state.all_classified:
            ct = case.actual_contact_type
            type_counts[ct] = type_counts.get(ct, 0) + 1

        lines.append("\nClassification Distribution:")
        for ct, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
            pct = count / state.total_cases * 100
            lines.append(f"  {ct}: {count} ({pct:.1f}%)")

    # Add patterns
    if state.patterns:
        lines.append(f"\nIdentified Patterns: {len(state.patterns)} clusters")
        for pattern in state.patterns[:5]:
            lines.append(f"  - {pattern.sub_theme}: {pattern.case_count} cases")

    # Add friction points
    if state.journey_map:
        lines.append(f"\nFriction Points: {len(state.journey_map)}")
        for friction in state.journey_map[:5]:
            lines.append(f"  - {friction.friction_point}: {friction.case_count} cases")

    # Add gaps
    if state.gap_table:
        critical_gaps = [g for g in state.gap_table if g.severity == 'Critical']
        lines.append(f"\nGaps Identified: {len(state.gap_table)} total, {len(critical_gaps)} critical")

    # Add FAQs
    if state.validated_faqs:
        lines.append(f"\nValidated FAQs: {len(state.validated_faqs)}")

    return "\n".join(lines)


def _create_basic_report_sections(state: PipelineState) -> None:
    """Create basic report structure without LLM — Arabic only."""
    state.report_sections_ar = {
        'executive_summary': {
            'heading': 'الملخص التنفيذي',
            'body': f'يحلل هذا التقرير {state.total_cases} استفسار وشكوى من المتعاملين. تشمل النتائج الرئيسية تحديد أنماط الخدمة، والنقاط الصعبة في رحلة المتعامل، والتوصيات لتحسين الخدمة.',
        },
        'methodology': {
            'heading': 'المنهجية',
            'body': 'تم إجراء التحليل باستخدام خط أنابيب متقدم يضم ستة مراحل: (1) التحقق من الصيغة، (2) التصنيف القائم على القواعد، (3) التصنيف المحسّن بالذكاء الاصطناعي، (4) تحليل الأنماط، (5) تحديد الفجوات، و (6) توليد التقرير.',
            'tables': [{
                'columns': ['المصدر', 'الطبيعة', 'الحجم', 'الفترة'],
                'rows': [
                    {
                        'المصدر': 'تحليل الاستفسارات',
                        'الطبيعة': 'بيانات CRM — نصوص غير مهيكلة',
                        'الحجم': f'{state.total_cases} حالة مغلقة',
                        'الفترة': convert_month_year_to_arabic(state.month_year) or 'يناير — مارس 2026'
                    },
                    {
                        'المصدر': 'دليل خدمات العملاء',
                        'الطبيعة': 'الدليل الرسمي يغطي خدمات المرور والترخيص والأمن',
                        'الحجم': '160 صفحة، 25 سؤالاً',
                        'الفترة': '2025'
                    }
                ],
                'row_count': 2,
                'col_count': 4,
                'original_index': 0
            }],
        },
        'classification_summary': {
            'heading': 'ملخص التصنيف',
            'body': 'تم تصنيف الحالات في ثماني فئات بناءً على نوايا المتعامل وأنماط التفاعل.',
        },
        'recommendations': {
            'heading': 'التوصيات',
            'body': 'بناءً على التحليل، يتم تقديم التوصيات التالية لتحسين جودة الخدمة ورضا المتعاملين.',
        },
    }


def run_stage6(
    state: PipelineState,
    excel_path: str,
    word_path: str,
    language: str = 'ar',
    api_key: str = ""
) -> PipelineState:
    """
    Stage 6: Artifact generator.

    Input: state with all_classified, analysis results, and report_sections
    Output: Excel and Word files on disk; report dictionary in state.report_json

    Args:
        state: Pipeline state from stages 1-5
        excel_path: Path to save Excel workbook
        word_path: Path to save Word document
        language: 'ar' or 'en'
        api_key: Anthropic API key for Word report generation

    Returns:
        Updated state with report_json dictionary
    """
    # Validation
    assert len(state.all_classified) == state.total_cases, "Case count mismatch"
    assert all(c.actual_contact_type for c in state.all_classified), "Missing contact types"
    assert all(0.0 <= c.confidence <= 1.0 for c in state.all_classified), "Invalid confidence scores"

    # FIX 1: Compute reclassification stats once — used by all downstream outputs
    reclassified = [
        c for c in state.all_classified
        if c.actual_contact_type != c.case_type
    ]
    state.reclassified_count = len(reclassified)
    state.reclassification_rate = (
        state.reclassified_count / state.total_cases * 100
        if state.total_cases > 0 else 0.0
    )

    # CRITICAL: Always clear and regenerate report sections to ensure LLM prompts
    # use the latest reconciled data (state.journey_map, state.gap_table, etc.)
    # and updated friction counts. Stale cached sections will have outdated prose.
    print(f"[Stage6] Clearing cached report_sections_ar and report_sections_en to ensure fresh generation...")
    state.report_sections_ar = {}
    state.report_sections_en = {}

    print(f"[Stage6] Calling _generate_report_sections with fresh state...")
    _generate_report_sections(state, api_key)
    print(f"[Stage6] After generation: {len(state.report_sections_ar or {})} sections in report_sections_ar")

    # Generate Excel
    generate_excel(state, excel_path)

    # Generate report dictionary first so state.report_json is always populated
    # (generate_word_report may be skipped when word_path is None)
    state.report_json = generate_json_report(state)

    # Generate Word report (skipped gracefully when word_path is None)
    generate_word_report(state, word_path, language, api_key)

    return state
