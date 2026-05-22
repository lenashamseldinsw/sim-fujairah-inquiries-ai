"""
STAGE 6: Artifact Generator (Complaints Flow)

Generates:
1. Excel workbook (openpyxl) — 9 sheets with complaint classification results
2. Word report (sword-word-builder) — 9-section bilingual report
3. Report dictionary — demo-compatible format for Streamlit display (in-memory)

Uses state.report_sections_ar and state.report_sections_en to build Word document and report dict.
Report dict is stored in state.report_json for passing to display functions.
"""

import json
import sys
import anthropic
from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import importlib.util
from concurrent.futures import ThreadPoolExecutor, as_completed

from .state import PipelineState, CaseRow, convert_month_year_to_arabic
from .stage6_json_report import generate_json_report
from .translate_report_en import translate_report_to_english
from .generate_workload_map_section import generate_workload_map_section
from .generate_customer_journey_section import generate_customer_journey_section
from .generate_digital_gaps_section import generate_digital_gaps_section
from .generate_digital_transformation_section import generate_digital_transformation_section
from .generate_ai_use_cases_section import generate_ai_use_cases_section
from .generate_improvement_roadmap_section import generate_improvement_roadmap_section
from .generate_conclusion_section import generate_conclusion_section

# Import build_report_ar to generate Arabic Word document from JSON
try:
    from .build_report_ar import build_report
except ImportError:
    build_report = None
    print("[Warning] Could not import build_report_ar")

# Import build_report_en to generate English Word document from JSON
try:
    from .build_report_en import build_report as build_report_en
except ImportError:
    build_report_en = None
    print("[Warning] Could not import build_report_en")


# Excel formatting
HEADER_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_ROW_FILL = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Complaints Excel sheet names (9 sheets, EXACT ORDER AND SPELLING)
COMPLAINT_SHEET_NAMES = [
    'ملخص التصنيف',           # Summary
    'جميع الحالات',           # All Cases
    'مكررة (مرفوضة)',         # Duplicate (Rejected)
    'بلا تصنيف خدمي',         # No Service Classification
    'خدمات مرورية',           # Traffic Services
    'أمنية وجنائية',          # Security & Criminal
    'شهادات وتصاريح',         # Certificates & Permits
    'خارج الاختصاص',          # Out of Scope
    'Pviot Table',            # KEEP TYPO
]

# Complaints Excel columns (21 columns, EXACT ORDER)
COMPLAINTS_EXCEL_COLUMNS = [
    'رقم_الطلب',
    'تفاصيل_الطلب',
    'الحل',
    'الخدمة',
    'الخدمة_الرئيسية',
    'شدة_الطلب',
    'نوع_المكالمة',
    'نوع_الشكوى',           # Pipeline OUTPUT classification
    'قناة_تقديم_الخدمة',
    'الحالة',
    'تاريخ_الإنشاء',
    'تاريخ_الإغلاق',
    'إغلاق_خلال_الوقت_المحدد',
    'تم_الحل_بواسطة',
    'الإدارة_العامة',
    'المالك',
    'اسم_مقدم_الطلب',
    'الجنسية',
    'رقم_الهوية',
    'الهاتف_الجوال',
    'الرقم_الوظيفي',
]

# Category sheet mapping for complaints
CATEGORY_SHEET_MAP = {
    'شكاوى مكررة (مرفوضة)':            'مكررة (مرفوضة)',
    'شكاوى بلا تصنيف خدمي ("أخرى")':  'بلا تصنيف خدمي',
    'شكاوى على الخدمات المرورية':       'خدمات مرورية',
    'شكاوى أمنية وجنائية':             'أمنية وجنائية',
    'شكاوى شهادات وتصاريح':            'شهادات وتصاريح',
    'شكاوى خارج الاختصاص والأخرى':     'خارج الاختصاص',
}


def _fix_taxonomy_consistency(cases: List[CaseRow]) -> List[CaseRow]:
    """Correct any sub_classification that is invalid for its top_level."""
    from .stage2_rules import ALL_COMPLAINT_SUB_CATEGORIES
    for case in cases:
        # All complaints have the same set of valid sub-categories
        valid_subs = ALL_COMPLAINT_SUB_CATEGORIES
        if case.sub_classification not in valid_subs:
            print(
                f"[Stage6] TAXONOMY FIX: case {case.case_number}: "
                f"sub '{case.sub_classification}' invalid, correcting to '{valid_subs[0]}'"
            )
            case.sub_classification = valid_subs[0]
    return cases


def generate_excel(state: PipelineState, output_path: str) -> None:
    """Generate Excel workbook with complaint classification results.

    Produces 9 sheets with complaint-specific categories:
    1. Summary
    2. All Cases
    3. Duplicate (Rejected)
    4. No Service Classification
    5. Traffic Services
    6. Security & Criminal
    7. Certificates & Permits
    8. Out of Scope
    9. Pivot Table
    """
    # Fix any cross-taxonomy mismatches before building any sheet
    state.all_classified = _fix_taxonomy_consistency(state.all_classified)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Summary
    ws_summary = wb.create_sheet('ملخص التصنيف', 0)
    _populate_summary_sheet(ws_summary, state)

    # Sheet 2: All Cases
    ws_all = wb.create_sheet('جميع الحالات', 1)
    _populate_all_cases_sheet(ws_all, state.all_classified, state)

    # Sheets 3–8: One per complaint sub-category, filtered and populated
    complaint_category_map = [
        ('شكاوى مكررة (مرفوضة)',            'مكررة (مرفوضة)'),
        ('شكاوى بلا تصنيف خدمي ("أخرى")',  'بلا تصنيف خدمي'),
        ('شكاوى على الخدمات المرورية',       'خدمات مرورية'),
        ('شكاوى أمنية وجنائية',             'أمنية وجنائية'),
        ('شكاوى شهادات وتصاريح',            'شهادات وتصاريح'),
        ('شكاوى خارج الاختصاص والأخرى',     'خارج الاختصاص'),
    ]

    for idx, (category_value, sheet_name) in enumerate(complaint_category_map, 2):
        subset = [c for c in state.all_classified if c.sub_classification == category_value]
        ws_category = wb.create_sheet(sheet_name, idx)
        _populate_all_cases_sheet(ws_category, subset, state)

    # Sheet 9: Pivot Table
    ws_pivot = wb.create_sheet('Pviot Table', 8)
    _populate_pivot_table_sheet(ws_pivot, state)

    wb.save(output_path)


def _populate_summary_sheet(ws, state: PipelineState) -> None:
    """Populate summary sheet for complaints with 6 complaint categories."""
    # Enable RTL layout
    ws.sheet_view.rightToLeft = True

    total = len(state.all_classified)

    ws.merge_cells('A1:D1')
    ws['A1'] = f"ملخص تحليل شكاوى شرطة الفجيرة — {convert_month_year_to_arabic(state.month_year) or 'Q1 2026'}"
    ws['A1'].font = Font(bold=True, size=14)

    ws.merge_cells('A2:D2')
    ws['A2'] = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d')}"

    # Count cases per complaint category
    complaint_counts = {}
    for case in state.all_classified:
        sub = case.sub_classification
        complaint_counts[sub] = complaint_counts.get(sub, 0) + 1

    # Category to sheet name mapping (matches complaint_category_map in artifact generation)
    category_to_sheet = {
        'شكاوى مكررة (مرفوضة)': 'مكررة (مرفوضة)',
        'شكاوى بلا تصنيف خدمي ("أخرى")': 'بلا تصنيف خدمي',
        'شكاوى على الخدمات المرورية': 'خدمات مرورية',
        'شكاوى أمنية وجنائية': 'أمنية وجنائية',
        'شكاوى شهادات وتصاريح': 'شهادات وتصاريح',
        'شكاوى خارج الاختصاص والأخرى': 'خارج الاختصاص',
    }

    # Extract counts for each category
    count_dup = complaint_counts.get('شكاوى مكررة (مرفوضة)', 0)
    count_other = complaint_counts.get('شكاوى بلا تصنيف خدمي ("أخرى")', 0)
    count_traffic = complaint_counts.get('شكاوى على الخدمات المرورية', 0)
    count_security = complaint_counts.get('شكاوى أمنية وجنائية', 0)
    count_certs = complaint_counts.get('شكاوى شهادات وتصاريح', 0)
    count_oor = complaint_counts.get('شكاوى خارج الاختصاص والأخرى', 0)

    # --- Section 1: Total cases ---
    row = 4
    ws[f'A{row}'] = "1. إجمالي الحالات"
    ws[f'A{row}'].font = Font(bold=True)
    row = 5
    ws[f'A{row}'] = "إجمالي الشكاوى المعالجة"
    ws[f'B{row}'] = total
    ws[f'C{row}'] = "100.0%"

    # --- Section 2: Distribution by complaint type ---
    row = 7
    ws[f'A{row}'] = "2. توزيع حسب نوع الشكوى"
    ws[f'A{row}'].font = Font(bold=True)
    row = 8

    summary_rows = [
        ('شكاوى مكررة (مرفوضة)',             count_dup,      'مكررة (مرفوضة)'),
        ('شكاوى بلا تصنيف خدمي ("أخرى")',   count_other,    'بلا تصنيف خدمي'),
        ('شكاوى على الخدمات المرورية',       count_traffic,  'خدمات مرورية'),
        ('شكاوى أمنية وجنائية',              count_security, 'أمنية وجنائية'),
        ('شكاوى شهادات وتصاريح',             count_certs,    'شهادات وتصاريح'),
        ('شكاوى خارج الاختصاص والأخرى',      count_oor,      'خارج الاختصاص'),
    ]

    totals_row = 5  # Row containing إجمالي الشكاوى
    for label, count, sheet_name in summary_rows:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"=IFERROR(B{row}/B{totals_row},0)"  # Live formula, not static %
        ws[f'D{row}'] = sheet_name  # Column D: sheet tab name
        ws[f'E{row}'] = None  # Column E: reserved for future notes (currently empty)
        row += 1

    # --- Section 3: Distribution by channel ---
    row += 1
    ws[f'A{row}'] = "3. توزيع حسب قناة التقديم"
    ws[f'A{row}'].font = Font(bold=True)
    channel_counts = {}
    for case in state.all_classified:
        ch = (case.case_channel or '').strip() or 'غير محدد'
        channel_counts[ch] = channel_counts.get(ch, 0) + 1
    row += 1
    for ch, count in sorted(channel_counts.items(), key=lambda x: x[1], reverse=True):
        ws[f'A{row}'] = ch
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"=IFERROR(B{row}/B{totals_row},0)"  # Live formula
        row += 1

    # --- Section 4: SLA on-time rate (formal rejection rate) ---
    row += 1
    ws[f'A{row}'] = "4. معدل الرفض الرسمي"
    ws[f'A{row}'].font = Font(bold=True)
    rejected = sum(1 for c in state.all_classified if c.case_status and c.case_status.strip() == 'طلب مرفوض')
    accepted = total - rejected
    row += 1
    accepted_row = row
    ws[f'A{row}'] = "شكاوى مقبولة"
    ws[f'B{row}'] = accepted
    ws[f'C{row}'] = f"=IFERROR(B{row}/B{totals_row},0)"  # Live formula
    row += 1
    rejected_row = row
    ws[f'A{row}'] = "شكاوى مرفوضة"
    ws[f'B{row}'] = rejected
    ws[f'C{row}'] = f"=IFERROR(B{row}/B{totals_row},0)"  # Live formula

    # --- Section 5: Proactive notifications (categories + cases) ---
    row += 2
    notification_categories = len(state.notification_opportunities or [])
    notification_cases = sum(
        n.get("cases_eliminated", 0) for n in (state.notification_opportunities or [])
    )
    ws[f'A{row}'] = f"5. الإشعارات الاستباقية ({notification_categories} فئة)"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "فئات الإشعار المكتشفة"
    ws[f'B{row}'] = notification_categories
    ws[f'C{row}'] = f"=IFERROR(B{row}/B{totals_row},0)"  # Live formula
    row += 1
    ws[f'A{row}'] = "حالات قابلة للإشعار الاستباقي"
    ws[f'B{row}'] = notification_cases
    ws[f'C{row}'] = f"=IFERROR(B{row}/B{totals_row},0)"  # Live formula

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20  # اسم الورقة (sheet name)
    ws.column_dimensions['E'].width = 15  # ملاحظات (remarks, currently empty)


def _populate_all_cases_sheet(ws, cases: List[CaseRow], state: PipelineState) -> None:
    """Populate sheet with complaint case data (RTL).

    Uses 21-column order for complaints:
    رقم_الطلب, تفاصيل_الطلب, الحل, الخدمة, الخدمة_الرئيسية, شدة_الطلب,
    نوع_المكالمة, نوع_الشكوى, قناة_تقديم_الخدمة, الحالة, تاريخ_الإنشاء,
    تاريخ_الإغلاق, إغلاق_خلال_الوقت_المحدد, تم_الحل_بواسطة, الإدارة_العامة,
    المالك, اسم_مقدم_الطلب, الجنسية, رقم_الهوية, الهاتف_الجوال, الرقم_الوظيفي
    """
    # Enable RTL layout
    ws.sheet_view.rightToLeft = True

    # Map normalized column names to CaseRow attributes
    NORMALIZED_TO_CASEROW = {
        'رقم_الطلب': 'case_number',
        'تفاصيل_الطلب': 'case_title',
        'تاريخ_الإنشاء': 'date_opened',
        'قناة_تقديم_الخدمة': 'case_channel',
        'نوع_المكالمة': 'case_type',
        'الخدمة_الرئيسية': 'service_name',
        'الحل': 'resolution_response',
        'الحالة': 'sla_color',
        'الإدارة_العامة': 'admin',
        'الخدمة': 'service',  # May not exist in CaseRow — pull from raw_df if available
    }

    # FIXED COLUMN ORDER (21 columns for complaints)
    headers = COMPLAINTS_EXCEL_COLUMNS

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Create a lookup dict for raw_df by case_number for columns not in CaseRow
    raw_df_lookup = {}
    if state.raw_df is not None:
        try:
            import pandas as pd
            df = state.raw_df
            # Get the normalized case number column
            case_num_col = 'رقم_الطلب'
            if case_num_col in df.columns:
                # Create lookup by case number
                for idx, row in df.iterrows():
                    case_num = str(row[case_num_col])
                    raw_df_lookup[case_num] = row
        except Exception as e:
            print(f"[Warning] Could not create raw_df lookup: {e}")

    # Write data rows
    for row_idx, case in enumerate(cases, 2):
        # Convert misclassification to نعم/لا
        reclassified = 'لا' if case.misclassification == 'OK' else 'نعم'

        # Get raw row data for this case (for unmapped columns)
        raw_row = raw_df_lookup.get(case.case_number)

        # Build data in the exact order specified by headers
        row_data = []

        for header_col in headers:
            value = ''

            # Input columns from CaseRow
            if header_col == 'رقم_الطلب':
                value = case.case_number
            elif header_col == 'تفاصيل_الطلب':
                value = case.case_title or ''
            elif header_col == 'الحل':
                value = case.resolution_response or ''
            elif header_col == 'الخدمة':
                # Try to get من raw_df (not in CaseRow)
                if raw_row is not None:
                    for col_name in ['الخدمة', 'الخدمة ', 'Service']:
                        if col_name in raw_row.index:
                            val = raw_row[col_name]
                            if val == val:  # pd.notna check
                                value = val
                            break
            elif header_col == 'الخدمة_الرئيسية':
                value = case.service_name or ''
            elif header_col == 'شدة_الطلب':
                # Extract from raw_df if available
                if raw_row is not None:
                    for col_name in ['شدة_الطلب', 'شدة الطلب', 'Severity']:
                        if col_name in raw_row.index:
                            val = raw_row[col_name]
                            if val == val:  # pd.notna check
                                value = val
                            break
            elif header_col == 'نوع_المكالمة':
                value = case.case_type or ''
            elif header_col == 'نوع_الشكوى':
                # Pipeline OUTPUT classification (sub_classification)
                value = case.sub_classification or ''
            elif header_col == 'قناة_تقديم_الخدمة':
                value = case.case_channel or ''
            elif header_col == 'الحالة':
                value = case.sla_color or ''
            elif header_col == 'تاريخ_الإنشاء':
                value = case.date_opened or ''
            elif header_col == 'تاريخ_الإغلاق':
                # Passthrough from raw_df — not stored in CaseRow
                if raw_row is not None:
                    for col_name in ['تاريخ_الإغلاق', 'تاريخ الإغلاق']:
                        if col_name in raw_row.index:
                            val = raw_row[col_name]
                            if val == val:  # pd.notna check
                                value = str(val) if str(val) not in ('nan', 'NaT', 'None') else ''
                            break
            elif header_col == 'إغلاق_خلال_الوقت_المحدد':
                # Extract from raw_df if available
                if raw_row is not None:
                    for col_name in ['إغلاق_خلال_الوقت_المحدد', 'إغلاق خلال الوقت المحدد']:
                        if col_name in raw_row.index:
                            val = raw_row[col_name]
                            if val == val:  # pd.notna check
                                value = val
                            break
            elif header_col == 'تم_الحل_بواسطة':
                # Extract from raw_df if available
                if raw_row is not None:
                    for col_name in ['تم_الحل_بواسطة', 'تم الحل بواسطة', 'Resolved By']:
                        if col_name in raw_row.index:
                            val = raw_row[col_name]
                            if val == val:  # pd.notna check
                                value = val
                            break
            elif header_col == 'الإدارة_العامة':
                value = case.admin or ''
            elif header_col == 'المالك':
                # Extract from raw_df if available
                if raw_row is not None:
                    for col_name in ['المالك', 'Owner']:
                        if col_name in raw_row.index:
                            val = raw_row[col_name]
                            if val == val:  # pd.notna check
                                value = val
                            break
            elif header_col == 'اسم_مقدم_الطلب':
                # Extract from raw_df if available
                if raw_row is not None:
                    for col_name in ['اسم_مقدم_الطلب', 'اسم مقدم الطلب', 'Submitter Name']:
                        if col_name in raw_row.index:
                            val = raw_row[col_name]
                            if val == val:  # pd.notna check
                                value = val
                            break
            elif header_col == 'الجنسية':
                # Extract from raw_df if available
                if raw_row is not None:
                    for col_name in ['الجنسية', 'Nationality']:
                        if col_name in raw_row.index:
                            val = raw_row[col_name]
                            if val == val:  # pd.notna check
                                value = val
                            break
            elif header_col == 'رقم_الهوية':
                # Extract from raw_df if available
                if raw_row is not None:
                    for col_name in ['رقم_الهوية', 'رقم الهوية', 'ID Number']:
                        if col_name in raw_row.index:
                            val = raw_row[col_name]
                            if val == val:  # pd.notna check
                                value = val
                            break
            elif header_col == 'الهاتف_الجوال':
                # Extract from raw_df if available
                if raw_row is not None:
                    for col_name in ['الهاتف_الجوال', 'الهاتف الجوال', 'Mobile']:
                        if col_name in raw_row.index:
                            val = raw_row[col_name]
                            if val == val:  # pd.notna check
                                value = val
                            break
            elif header_col == 'الرقم_الوظيفي':
                # Extract from raw_df if available
                if raw_row is not None:
                    for col_name in ['الرقم_الوظيفي', 'الرقم الوظيفي', 'Employee ID']:
                        if col_name in raw_row.index:
                            val = raw_row[col_name]
                            if val == val:  # pd.notna check
                                value = val
                            break

            row_data.append(value)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL
            cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)

    # Auto-size columns (fixed order per headers)
    ws.column_dimensions['A'].width = 12   # رقم_الطلب
    ws.column_dimensions['B'].width = 80   # تفاصيل_الطلب
    ws.column_dimensions['C'].width = 60   # الحل
    ws.column_dimensions['D'].width = 20   # الخدمة
    ws.column_dimensions['E'].width = 20   # الخدمة_الرئيسية
    ws.column_dimensions['F'].width = 15   # شدة_الطلب
    ws.column_dimensions['G'].width = 15   # نوع_المكالمة
    ws.column_dimensions['H'].width = 18   # نوع_الشكوى
    ws.column_dimensions['I'].width = 18   # قناة_تقديم_الخدمة
    ws.column_dimensions['J'].width = 15   # الحالة
    ws.column_dimensions['K'].width = 18   # تاريخ_الإنشاء
    ws.column_dimensions['L'].width = 18   # تاريخ_الإغلاق
    ws.column_dimensions['M'].width = 18   # إغلاق_خلال_الوقت_المحدد
    ws.column_dimensions['N'].width = 18   # تم_الحل_بواسطة
    ws.column_dimensions['O'].width = 18   # الإدارة_العامة
    ws.column_dimensions['P'].width = 15   # المالك
    ws.column_dimensions['Q'].width = 18   # اسم_مقدم_الطلب
    ws.column_dimensions['R'].width = 15   # الجنسية
    ws.column_dimensions['S'].width = 15   # رقم_الهوية
    ws.column_dimensions['T'].width = 15   # الهاتف_الجوال
    ws.column_dimensions['U'].width = 15   # الرقم_الوظيفي

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(cases) + 1}'


def _populate_pivot_table_sheet(ws, state: PipelineState) -> None:
    """Populate pivot table sheet.

    Column A: unique الخدمة values from raw input
    Column B: count of cases for each service
    Columns D-E: mirror columns A-B using cell references
    """
    ws.sheet_view.rightToLeft = True

    # Extract unique services and their counts
    service_counts = {}
    if state.raw_df is not None:
        try:
            import pandas as pd
            df = state.raw_df
            service_col = 'الخدمة'
            if service_col in df.columns:
                for idx, row in df.iterrows():
                    service = str(row[service_col]).strip()
                    if service and service != 'nan':
                        service_counts[service] = service_counts.get(service, 0) + 1
        except Exception as e:
            print(f"[Warning] Could not extract service counts: {e}")

    # Write headers
    ws['A1'] = 'الخدمة'
    ws['B1'] = 'العدد'
    ws['D1'] = 'الخدمة'
    ws['E1'] = 'العدد'

    # Format headers
    for cell in ['A1', 'B1', 'D1', 'E1']:
        ws[cell].fill = HEADER_FILL
        ws[cell].font = HEADER_FONT
        ws[cell].border = BORDER

    # Write data
    row = 2
    for service, count in sorted(service_counts.items(), key=lambda x: x[1], reverse=True):
        ws[f'A{row}'] = service
        ws[f'B{row}'] = count
        ws[f'D{row}'] = f'=A{row}'
        ws[f'E{row}'] = f'=B{row}'

        for col in ['A', 'B', 'D', 'E']:
            ws[f'{col}{row}'].border = BORDER
            if row % 2 == 0:
                ws[f'{col}{row}'].fill = ALT_ROW_FILL

        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12


def generate_word_report(
    state: PipelineState,
    output_path: str,
    language: str = 'ar',
    api_key: str = ""
) -> None:
    """
    Generate Arabic and English Word reports from the pipeline state.

    Generates the JSON report, saves it to disk, then:
    - Uses build_report_ar to produce a styled Arabic RTL .docx
    - Uses build_report_en to produce a styled English LTR .docx (when the
      English translation is available in state.report_json_en)

    Args:
        state: Pipeline state with report_sections
        output_path: Base path for .docx output (Arabic). English docx is saved
                     alongside it with an ``_en`` suffix before the extension.
        language: Kept for backward compatibility.
        api_key: Anthropic API key for LLM report generation.
    """
    if output_path is None:
        print("⚠️  Skipping Word report generation (no output_path provided)")
        return

    # Generate JSON report from state
    report_data = generate_json_report(state)

    # Create output path and derive JSON path from it
    output_path = Path(output_path)
    json_path = output_path.with_stem(output_path.stem + "_data").with_suffix(".json")

    # Save Arabic JSON to disk
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(report_data, f, ensure_ascii=False, indent=2)
    print(f"[Stage6] Saved report JSON: {json_path}")

    # Save English JSON alongside Arabic if translation succeeded
    en_json_path = None
    if getattr(state, 'report_json_en', None):
        en_json_path = output_path.with_stem(output_path.stem + "_data_en").with_suffix(".json")
        with open(en_json_path, 'w', encoding='utf-8') as f:
            json.dump(state.report_json_en, f, ensure_ascii=False, indent=2)
        print(f"[Stage6] Saved English report JSON: {en_json_path}")

    # Build Arabic Word document from JSON
    if build_report is not None:
        build_report(json_path, output_path)
        print(f"[Stage6] Generated Arabic Word report: {output_path}")
    else:
        print("⚠️  Skipping Arabic Word report generation (build_report_ar not available)")

    # Build English Word document from English JSON
    if build_report_en is not None and en_json_path is not None:
        en_docx_path = output_path.with_stem(output_path.stem + "_en")
        build_report_en(en_json_path, en_docx_path)
        print(f"[Stage6] Generated English Word report: {en_docx_path}")
    elif build_report_en is None:
        print("⚠️  Skipping English Word report generation (build_report_en not available)")
    else:
        print("⚠️  Skipping English Word report generation (English JSON translation unavailable)")


def _generate_report_sections(state: PipelineState, api_key: str = "") -> None:
    """
    Generate report sections via LLM with threading.

    Uses 3-wave parallelization:
    - Wave 1: 7 independent sections in parallel (executive_summary, methodology, workload_map,
              customer_journey, digital_gaps, digital_transformation, ai_use_cases)
    - Wave 2: improvement_roadmap (depends on ai_use_cases from wave 1)
    - Wave 3: conclusion (depends on both ai_use_cases and improvement_roadmap from earlier waves)

    Thread safety: Each generator function is read-only on state and returns a result dict.
    Results are collected in a local dict per thread, then merged into state.report_sections_ar
    after all threads complete (single-threaded merge, no race conditions).
    """
    print(f"[GenSections] api_key present: {bool(api_key)}")
    print(f"[GenSections] api_key length: {len(api_key) if api_key else 0}")

    if not api_key:
        raise ValueError("API key is required to generate report sections")

    # Initialize Arabic-only dict
    state.report_sections_ar = {}

    # ──────────────────────────────────────────────────────────────────────────────────
    # WAVE 1: 7 independent sections — run in parallel
    # ──────────────────────────────────────────────────────────────────────────────────
    print("[GenSections] WAVE 1: Starting 7 parallel section generations...")

    wave1_tasks = {
        'executive_summary': (generate_executive_summary_section, 'أولاً: الملخص التنفيذي — التحليلات الرئيسية'),
        'methodology': (generate_methodology_section, 'ثانياً: المنهجية وطبيعة المصادر'),
        'workload_map': (generate_workload_map_section, 'ثالثاً: التحليل الأول — خريطة تصنيف الطلبات'),
    }

    # digital_gaps requires gap_table from Stage 5; skip gracefully if unavailable
    if state.gap_table:
        wave1_tasks['digital_gaps'] = (generate_digital_gaps_section, 'خامساً: التحليل الثالث — تحليل الفجوات الرقمية')
    else:
        print("[GenSections] WARNING: state.gap_table is empty — skipping digital_gaps section")

    # digital_transformation section — always include to maintain section numbering
    # Will be generated even if FAQ data is minimal; LLM will provide context-driven content
    wave1_tasks['digital_transformation'] = (generate_digital_transformation_section, 'سادساً: التحليل الرابع — خطة التحويل الرقمي')
    if not (state.validated_faqs or state.faq_candidates):
        print("[GenSections] NOTE: digital_transformation will be generated without FAQ data from stages 4/5")

    # ai_use_cases requires gap_table from Stage 5; skip gracefully if unavailable
    if state.gap_table:
        wave1_tasks['ai_use_cases'] = (generate_ai_use_cases_section, 'سابعاً: حالات الاستخدام المدعومة بالذكاء الاصطناعي')
    else:
        print("[GenSections] WARNING: state.gap_table is empty — skipping ai_use_cases section")

    # customer_journey requires journey_map from Stage 4; skip gracefully if unavailable
    if state.journey_map:
        wave1_tasks['customer_journey'] = (
            generate_customer_journey_section,
            'رابعاً: التحليل الثاني — التحديات في رحلة المتعامل',
        )
    else:
        print("[GenSections] WARNING: state.journey_map is empty — skipping customer_journey section")

    wave1_results = {}  # Local dict — no shared state touched inside threads

    def run_section(key: str, fn, state: PipelineState, api_key: str) -> tuple[str, Dict[str, Any]]:
        """Run a section generator with up to 3 retries on failure."""
        max_attempts = 3
        last_error: Exception = RuntimeError(f"[GenSections] {key}: no attempts made")
        for attempt in range(1, max_attempts + 1):
            try:
                if attempt > 1:
                    print(f"[GenSections] Retrying {key} (attempt {attempt}/{max_attempts})...")
                else:
                    print(f"[GenSections] Starting {key}...")
                result = fn(state, api_key)  # Pure: reads state, returns dict
                print(f"[GenSections] ✓ {key} complete (attempt {attempt})")
                return key, result
            except Exception as e:
                last_error = e
                print(f"[GenSections] ✗ {key} failed on attempt {attempt}: {e}")
        raise last_error

    with ThreadPoolExecutor(max_workers=7) as executor:
        futures = {
            executor.submit(run_section, key, fn, state, api_key): key
            for key, (fn, _heading) in wave1_tasks.items()
        }
        for future in as_completed(futures):
            key = futures[future]
            try:
                result_key, result = future.result()
                wave1_results[result_key] = result
            except Exception as e:
                raise RuntimeError(f"[GenSections] Wave 1 failed on '{key}': {e}") from e

    print(f"[GenSections] ✓ WAVE 1 complete: {len(wave1_results)} sections generated")

    # ──────────────────────────────────────────────────────────────────────────────────
    # Merge Wave 1 results into state (single-threaded, no race conditions)
    # ──────────────────────────────────────────────────────────────────────────────────
    print("[GenSections] Merging WAVE 1 results into state...")

    # Special handling for executive_summary and methodology (have special structure)
    if 'executive_summary' in wave1_results:
        exec_summary = wave1_results['executive_summary']
        state.report_sections_ar['executive_summary'] = {
            'heading': wave1_tasks['executive_summary'][1],
            'body': exec_summary.get('framing_paragraph', ''),
            'tables': [exec_summary.get('key_findings', [])],
            'core_message': exec_summary.get('core_message', ''),
            'raw_data': exec_summary
        }

    if 'methodology' in wave1_results:
        methodology = wave1_results['methodology']
        sources_rows = methodology.get('sources_table', [])
        sources_ar = {
            'columns': ['المصدر', 'الطبيعة', 'الحجم', 'الفترة'],
            'rows': sources_rows,
            'row_count': len(sources_rows),
            'col_count': 4,
        }
        if not _is_valid_table(sources_ar):
            raise RuntimeError("[GenSections] Arabic sources table is invalid or empty")
        state.report_sections_ar['methodology'] = {
            'heading': wave1_tasks['methodology'][1],
            'classification_method': methodology.get('classification_method', ''),
            'analyzed_fields': methodology.get('analyzed_fields', ''),
            'tables': [sources_ar],
            'raw_data': methodology
        }

    # Standard sections (workload_map, customer_journey, digital_gaps, digital_transformation, ai_use_cases)
    standard_sections = ['workload_map', 'customer_journey', 'digital_gaps', 'digital_transformation', 'ai_use_cases']
    for section_key in standard_sections:
        if section_key in wave1_results:
            state.report_sections_ar[section_key] = {
                'heading': wave1_tasks[section_key][1],
                'raw_data': wave1_results[section_key],
            }

    print("[GenSections] ✓ Wave 1 results merged")

    # ──────────────────────────────────────────────────────────────────────────────────
    # WAVE 2: improvement_roadmap (depends on ai_use_cases from Wave 1)
    # ──────────────────────────────────────────────────────────────────────────────────
    print("[GenSections] WAVE 2: Generating Improvement Roadmap (depends on AI Use Cases)...")
    roadmap = generate_improvement_roadmap_section(state, api_key)
    state.report_sections_ar['improvement_roadmap'] = {
        'heading': 'ثامناً: خارطة الطريق التحسينية المقترحة',
        'raw_data': roadmap,
    }
    print("[GenSections] ✓ WAVE 2 complete")

    # ──────────────────────────────────────────────────────────────────────────────────
    # WAVE 3: conclusion (depends on both ai_use_cases and improvement_roadmap)
    # ──────────────────────────────────────────────────────────────────────────────────
    print("[GenSections] WAVE 3: Generating Conclusion (depends on AI Use Cases + Roadmap)...")
    conclusion = generate_conclusion_section(state, api_key)
    state.report_sections_ar['conclusion'] = {
        'heading': 'تاسعاً: الخلاصة — من البيانات إلى القرار',
        'raw_data': conclusion,
    }
    print("[GenSections] ✓ WAVE 3 complete")

    print("[Report Gen] ✅ All report sections generated successfully")


def _fix_unescaped_newlines(json_str: str) -> str:
    """Fix unescaped newlines in JSON string values (from LLM responses).

    When Claude returns JSON with literal newlines in Arabic text,
    this escapes them properly so json.loads() can parse it.
    """
    result = []
    i = 0
    in_string = False
    escape_next = False

    while i < len(json_str):
        char = json_str[i]

        if escape_next:
            result.append(char)
            escape_next = False
            i += 1
            continue

        if char == '\\' and in_string:
            result.append(char)
            escape_next = True
            i += 1
            continue

        if char == '"':
            in_string = not in_string
            result.append(char)
            i += 1
            continue

        if char == '\n' and in_string:
            # Replace literal newline with escaped version
            result.append('\\n')
            i += 1
            continue

        result.append(char)
        i += 1

    return ''.join(result)


def _build_pre_computed_findings(
    total_cases: int,
    misclassification_count: int,
    misclassification_rate: float,
    dominant_type: str,
    dominant_type_count: int,
    dominant_type_pct: float,
    complaint_subcategories: list,
    friction_points: list,
    friction_count: int,
    sla_closed: int,
    sla_rate: float,
) -> list:
    """
    Pre-compute the findings table deterministically from state data.

    This avoids the ambiguity where the LLM combined case counts from multiple friction points
    and readers couldn't distinguish between "N friction points" and "N cases".

    Each finding row is built from a template where:
    - The title (الاكتشاف) is ALWAYS deterministic from state, never LLM-generated
    - Case counts are separated from friction point counts in the description
    - Friction point groupings are explicit: "نقطتا احتكاك" (2 points) vs "N حالة" (N cases)

    Returns list of dicts with keys: number, title, description, importance
    """
    findings = []

    # ROW 1 — Classification accuracy gap
    findings.append({
        "number": 1,
        "title": f"تصنيف غير دقيق بنسبة {misclassification_rate:.1f}%",
        "description": (
            f"كانت {misclassification_count} من {total_cases} حالة مُصنَّفة أصلاً بشكل غير صحيح. "
            f"تمثل هذه الفجوة أساس التحديات التشغيلية المُكتشفة في هذا التحليل."
        ),
        "importance": "🔴 حرجة"
    })

    # ROW 2 — Dominant complaint type dominance
    top_2_complaints = complaint_subcategories[:2]
    complaint_text = ""
    if len(top_2_complaints) >= 2:
        complaint_text = (
            f"أكبرها: {top_2_complaints[0]['name']} ({top_2_complaints[0]['count']} حالة) و"
            f"{top_2_complaints[1]['name']} ({top_2_complaints[1]['count']} حالة)"
        )
    elif len(top_2_complaints) == 1:
        complaint_text = f"{top_2_complaints[0]['name']} ({top_2_complaints[0]['count']} حالة)"

    findings.append({
        "number": 2,
        "title": f"الشكاوى تهيمن بـ {dominant_type_pct:.1f}% على عبء العمل",
        "description": (
            f"{dominant_type_count} حالة من {total_cases} كانت شكاوى بعد إعادة التصنيف. "
            f"{complaint_text}. هذا التركيز يعكس محور التحسين الأساسي للعمليات."
        ),
        "importance": "🔴 حرجة"
    })

    # ROW 3 — Largest friction point (SINGLE POINT, not grouped)
    if friction_points:
        top_friction = friction_points[0]
        findings.append({
            "number": 3,
            "title": f"نقطة احتكاك واحدة: {top_friction['name']} ({top_friction['case_count']} حالة، {round(top_friction['case_count']/total_cases*100, 1)}%)",
            "description": (
                f"تؤثر هذه النقطة على {top_friction['case_count']} حالة ({round(top_friction['case_count']/total_cases*100, 1)}% من الإجمالي). "
                f"السبب الجذري: {_root_cause_label(top_friction['root_cause'])}"
            ),
            "importance": top_friction['gap_severity']
        })

    # ROW 4 — Multiple friction points grouped by shared root cause
    # FIX: Use friction point COUNT, not case count, as the leading number
    if len(friction_points) >= 2:
        grouped_by_cause = {}
        for fp in friction_points[1:]:
            cause = fp['root_cause']
            if cause not in grouped_by_cause:
                grouped_by_cause[cause] = []
            grouped_by_cause[cause].append(fp)

        # Pick the largest group
        largest_group = max(grouped_by_cause.values(), key=lambda g: sum(f['case_count'] for f in g))
        friction_point_count_in_group = len(largest_group)
        case_count_in_group = sum(f['case_count'] for f in largest_group)
        group_cause = largest_group[0]['root_cause']
        group_cause_label = _root_cause_label(group_cause)

        # Title: Leading number is FRICTION POINT COUNT, not case count
        if friction_point_count_in_group == 1:
            point_label = "نقطة احتكاك واحدة"
        elif friction_point_count_in_group == 2:
            point_label = "نقطتا احتكاك"
        else:
            point_label = f"{friction_point_count_in_group} نقاط احتكاك"

        # Build point list for description
        point_names = ", ".join([f['name'] for f in largest_group])

        findings.append({
            "number": 4,
            "title": f"{point_label} — {group_cause_label} ({case_count_in_group} حالة)",
            "description": (
                f"تشترك {point_label} في سبب جذري مشترك: {point_names}. "
                f"إجمالي الحالات المتأثرة: {case_count_in_group} حالة. "
                f"تحسين هذه المجموعة سيسهم بشكل كبير في تقليل الاحتكاكات."
            ),
            "importance": largest_group[0]['gap_severity']
        })

    # ROW 5 — SLA / operational performance
    findings.append({
        "number": 5,
        "title": f"{sla_rate:.1f}% قبول في الوقت المحدد",
        "description": (
            f"تم قبول {sla_closed} من {total_cases} حالة في الوقت المحدد (معدل {sla_rate:.1f}%). "
            f"هذا يثبت القدرة التشغيلية، لكنه لا يعالج الفجوات الهيكلية في الدقة والفهم."
        ),
        "importance": "🟢 إيجابية"
    })

    return findings


def _root_cause_label(root_cause_category: str) -> str:
    """Map root cause category to Arabic label."""
    mapping = {
        'missing_info': 'غياب معلومات من الدليل',
        'inaccessible_info': 'معلومات موجودة لكنها صعبة الوصول',
        'no_proactive_notification': 'غياب الإشعار الاستباقي',
        'platform_bug': 'خلل تقني في المنصة',
        'policy_complexity': 'تعقيد إجراءات السياسة'
    }
    return mapping.get(root_cause_category, root_cause_category)


def generate_executive_summary_section(state: PipelineState, api_key: str) -> Dict[str, Any]:
    """
    Generate the executive summary section using Claude API.

    Extracts all necessary metrics from state and calls Claude with the detailed prompt.

    Args:
        state: Pipeline state with classified cases and analysis results
        api_key: Anthropic API key

    Returns:
        Dict with section_key and content following the specified JSON structure
    """
    try:
        # Extract classification metrics
        total_cases = state.total_cases
        all_classified = state.all_classified or []

        # Calculate distribution (after reclassification)
        distribution = {}
        original_distribution = {}

        for case in all_classified:
            final_type = case.top_level  # Final classification (after Stage 2 + Stage 3)
            distribution[final_type] = distribution.get(final_type, 0) + 1

            # Track misclassifications against original CRM label
            original_type = case.case_type  # Original CRM نوع_المكالمة label
            original_distribution[original_type] = original_distribution.get(original_type, 0) + 1

        misclassification_count = state.reclassified_count
        misclassification_rate = state.reclassification_rate
        matched_original_count = total_cases - misclassification_count
        matched_original_rate = (matched_original_count / total_cases * 100) if total_cases > 0 else 0

        # Find dominant contact type
        dominant_type = max(distribution.items(), key=lambda x: x[1])[0] if distribution else ""
        dominant_type_count = distribution.get(dominant_type, 0)
        dominant_type_pct = (dominant_type_count / total_cases * 100) if total_cases > 0 else 0

        # Extract complaint subcategories from patterns
        complaint_subcategories = []
        if state.patterns:
            complaint_patterns = [p for p in state.patterns if 'شكوى' in (p.cluster_ar or p.cluster)]
            complaint_patterns.sort(key=lambda x: x.case_count, reverse=True)
            for pattern in complaint_patterns[:10]:
                pct = (pattern.case_count / dominant_type_count * 100) if dominant_type_count > 0 else 0
                complaint_subcategories.append({
                    'name': pattern.sub_theme_ar or pattern.sub_theme,
                    'count': pattern.case_count,
                    'pct_of_complaints': pct
                })

        # Extract friction points
        friction_points = []
        top_friction_point_name = ""
        top_friction_case_count = 0
        top_friction_pct = 0.0

        if state.journey_map:
            for friction in state.journey_map[:10]:
                # Map severity from root_cause_category
                severity_map = {
                    'missing_info': '🔴 حرجة',
                    'inaccessible_info': '🔴 حرجة',
                    'no_proactive_notification': '🟡 عالية',
                    'platform_bug': '🔴 حرجة',
                    'policy_complexity': '🟡 عالية'
                }
                severity = severity_map.get(friction.root_cause_category, '🟢 متوسطة')

                friction_points.append({
                    'name': friction.friction_point_ar or friction.friction_point,
                    'case_count': friction.case_count,
                    'root_cause': friction.root_cause_category,
                    'gap_severity': severity
                })

            # Extract top friction for direct injection into prompt
            if friction_points:
                top_friction_point_name = friction_points[0]['name']
                top_friction_case_count = friction_points[0]['case_count']
                top_friction_pct = round(top_friction_case_count / total_cases * 100, 1) if total_cases > 0 else 0.0

        # Calculate SLA metrics — check for 'نعم' (yes) in SLA compliance field
        # Must be done BEFORE _build_pre_computed_findings call
        sla_closed = sum(1 for c in all_classified if c.sla_color == 'نعم')
        sla_rate = (sla_closed / total_cases * 100) if total_cases > 0 else 0

        # FIX: Pre-compute findings table deterministically from state before LLM call
        # Avoids LLM inventing ambiguous case-count titles
        friction_count = len(state.journey_map or [])
        pre_computed_findings = _build_pre_computed_findings(
            total_cases=total_cases,
            misclassification_count=misclassification_count,
            misclassification_rate=misclassification_rate,
            dominant_type=dominant_type,
            dominant_type_count=dominant_type_count,
            dominant_type_pct=dominant_type_pct,
            complaint_subcategories=complaint_subcategories,
            friction_points=friction_points,
            friction_count=friction_count,
            sla_closed=sla_closed,
            sla_rate=sla_rate,
        )

        # Extract gaps with enhanced guidebook intelligence
        gap_table = []
        guidebook_coverage_metrics = {
            'total_gaps': 0,
            'covered': 0,
            'partially_covered': 0,
            'missing': 0,
            'avg_coverage_percentage': 0,
            'avg_match_confidence': 0,
            'gaps_with_proactive_notification_opportunity': 0
        }

        if state.gap_table:
            coverage_percentages = []
            match_confidences = []

            for gap in state.gap_table[:20]:
                severity_emoji = '🔴 حرجة' if gap.severity == 'Critical' else '🟡 عالية' if gap.severity == 'Medium' else '🟢 متوسطة'

                # Track coverage metrics
                guidebook_coverage_metrics['total_gaps'] += 1
                if gap.guidebook_status == 'Covered':
                    guidebook_coverage_metrics['covered'] += 1
                elif gap.guidebook_status == 'Partially Covered':
                    guidebook_coverage_metrics['partially_covered'] += 1
                else:
                    guidebook_coverage_metrics['missing'] += 1

                if gap.coverage_percentage:
                    coverage_percentages.append(gap.coverage_percentage)
                if gap.guidebook_match_confidence:
                    match_confidences.append(gap.guidebook_match_confidence)
                if gap.proactive_notification_opportunity:
                    guidebook_coverage_metrics['gaps_with_proactive_notification_opportunity'] += 1

                gap_table.append({
                    'topic': gap.topic_ar or gap.topic,
                    'case_count': gap.case_count,
                    'current_status': gap.guidebook_status,
                    'gap_type': gap.gap_type_ar or gap.gap_type,
                    'coverage_percentage': gap.coverage_percentage,
                    'clarity': gap.clarity_assessment,
                    'format': gap.format_assessment,
                    'has_visuals': gap.has_visual_guidance,
                    'match_confidence': gap.guidebook_match_confidence,
                    'can_use_proactive_notification': gap.proactive_notification_opportunity,
                    'guidebook_excerpt': gap.guidebook_excerpt_ar or gap.guidebook_excerpt,
                    'recommendation': gap.recommendation
                })

            # Calculate averages
            if coverage_percentages:
                guidebook_coverage_metrics['avg_coverage_percentage'] = sum(coverage_percentages) / len(coverage_percentages)
            if match_confidences:
                guidebook_coverage_metrics['avg_match_confidence'] = sum(match_confidences) / len(match_confidences)

        # Calculate proactive notification impact (from notification opportunities)
        proactive_notification_total = 0
        if state.notification_opportunities:
            proactive_notification_total = sum(n.get('cases_eliminated', n.get('case_count', 0)) for n in state.notification_opportunities)
        proactive_notification_pct = (proactive_notification_total / total_cases * 100) if total_cases > 0 else 0

        # Digital channel percentage
        digital_cases = sum(1 for c in all_classified if c.case_channel in ['app', 'web', 'website', 'application'])
        digital_channel_pct = (digital_cases / total_cases * 100) if total_cases > 0 else 0

        # Date range extraction
        date_range = convert_month_year_to_arabic(state.month_year) or "يناير — مارس 2026"
        quarter_label = "Q1 2026"

        # Serialize pre-computed findings for the prompt
        findings_json = json.dumps(pre_computed_findings, ensure_ascii=False, indent=2)

        # Build the prompt with PRE-COMPUTED findings table (locked, not LLM-generated)
        prompt = f"""You are an expert CX strategist writing the executive summary of a formal Arabic government report on complaint analysis.

OUTPUT LANGUAGE: Arabic only. Do not generate English translations.

INPUTS PROVIDED:
- total_cases: {total_cases}
- date_range: {date_range}
- friction_count: {friction_count} (total distinct friction points identified)
- gap_table: {json.dumps(gap_table, ensure_ascii=False)}
- guidebook_coverage_metrics: {json.dumps(guidebook_coverage_metrics, ensure_ascii=False)}
- sla_rate: {sla_rate:.1f}%

KEY STRUCTURAL INSIGHT FOR YOUR REFERENCE:
The main reclassification finding: {misclassification_rate:.1f}% of cases were initially misclassified.
When corrected, {dominant_type} rises to {dominant_type_pct:.1f}% of total workload ({dominant_type_count} cases).

YOUR TASK — WRITE ONLY TWO SECTIONS:
─────────────────────────────────────────────

1. FRAMING PARAGRAPH (فقرة الإطار)
   Write 2–3 sentences that:
   - State how many complaint cases were analyzed, from which CRM source, and for which period
   - Name both data sources: "تصدير نظام إدارة علاقات العملاء" and "دليل الخدمات والأسئلة الشائعة"
   - Declare the report's purpose: transforming data into actionable decisions, not just presenting numbers
   - End with the most striking structural discovery: the {dominant_type} reclassification finding
   Style: Open with "يُقدّم هذا التقرير...". Third sentence must begin with
   "المُستجد الجوهري:" and deliver insight, not description.

2. الرسالة الجوهرية (CORE MESSAGE)
   One paragraph placed AFTER the locked findings table below. Must:
   - Begin with "الرسالة الجوهرية:"
   - Name the specific systemic failure: {misclassification_rate:.1f}% misclassification + {friction_count} distinct friction points
   - State what fixing it unlocks, using specific numbers from the inputs
   - End with a forward-looking statement about data-driven strategy

   Style: 3 sentences maximum. No bullet points. Assertive, not descriptive.

─────────────────────────────────────────────
LOCKED FINDINGS TABLE — COPY VERBATIM, DO NOT MODIFY:
─────────────────────────────────────────────

This table is LOCKED and pre-computed from authoritative state data.
Your ONLY task is to write framing_paragraph and core_message.
Do NOT invent, modify, or re-order the findings table rows.
Do NOT change any number in the table — case counts, percentages, or friction point descriptions.

Pre-computed key_findings table (5 rows):
{findings_json}

CRITICAL DISTINCTION (重要):
- Friction point COUNT = total number of distinct access barriers (نقاط احتكاك)
- Case COUNT = number of cases affected by that friction point (حالات)
- These are different numbers. Do NOT confuse them.
- Example: If ROW 4 says "نقطتا احتكاك — ... (10 حالة)", that means:
  - 2 friction points (نقطتا احتكاك)
  - Affecting 10 cases total (10 حالة)
  - NOT 10 friction points

VALIDATION RULE:
If any finding headline number exceeds {friction_count}, that finding is invalid.
Example: "10 نقاط احتكاك" is invalid if friction_count={friction_count}.
(This validates that we're counting the right dimension.)

─────────────────────────────────────────────
OUTPUT FORMAT — return JSON with THREE fields:
─────────────────────────────────────────────
{{
  "section": "executive_summary",
  "framing_paragraph": "...",  ← Your writing (2–3 sentences, Arabic only)
  "key_findings": {findings_json},  ← LOCKED: return as provided above, no changes
  "core_message": "..."  ← Your writing (1 paragraph, 3 sentences max)
}}

RULES:
- key_findings: Return EXACTLY as provided above — every row, every value, unchanged.
- framing_paragraph: Arabic only, 2–3 sentences, cite date_range and both data sources.
- core_message: Arabic only, 3 sentences max, begin with "الرسالة الجوهرية:".
- No markdown, no extra keys, no extra nesting.
- CRITICAL: Do NOT use double-quote characters (") inside any string value. Use « » for citations.
"""

        client = anthropic.Anthropic(api_key=api_key)
        print(f"[ExecSummary] Calling API with model claude-sonnet-4-6")
        print(f"[ExecSummary] total_cases={total_cases}, reclassified={misclassification_count}")
        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4000,
            messages=[
                {
                    "role": "user",
                    "content": prompt
                }
            ]
        )

        # Extract JSON from response
        response_text = message.content[0].text

        # Try to parse JSON from response
        import re

        # First try: look for ```json ... ``` block (with or without closing ```)
        result = None
        json_code_block = re.search(r'```\s*(?:json)?\s*\n(.*?)(?:\n```|$)', response_text, re.DOTALL)
        if json_code_block:
            json_candidate = json_code_block.group(1).strip()
            try:
                result = json.loads(json_candidate)
            except json.JSONDecodeError:
                pass

        # Second try: extract between first { and last } using state machine
        if not result:
            first_brace = response_text.find('{')
            if first_brace != -1:
                depth = 0
                in_string = False
                escape = False

                for i in range(first_brace, len(response_text)):
                    char = response_text[i]

                    if escape:
                        escape = False
                        continue

                    if char == '\\' and in_string:
                        escape = True
                        continue

                    if char == '"':
                        in_string = not in_string
                        continue

                    if in_string:
                        continue

                    if char == '{':
                        depth += 1
                    elif char == '}':
                        depth -= 1
                        if depth == 0:
                            json_str = response_text[first_brace:i + 1]
                            try:
                                result = json.loads(json_str)
                                break
                            except json.JSONDecodeError as e:
                                fixed_json = _fix_unescaped_newlines(json_str)
                                try:
                                    result = json.loads(fixed_json)
                                    break
                                except json.JSONDecodeError:
                                    raise RuntimeError(
                                        f"[ExecSummary] Failed to parse executive summary JSON: {e}\n"
                                        f"Attempted string length: {len(json_str)}\n"
                                        f"First 500 chars: {json_str[:500]}\n"
                                        f"Last 500 chars: {json_str[-500:] if len(json_str) > 500 else 'N/A'}"
                                    )

        if not result:
            print("No JSON found in executive summary response")
            print(f"Response first 500 chars: {response_text[:500]}")
            raise RuntimeError("Executive summary: No JSON found in API response")

        # Reinject pre-computed findings
        result['key_findings'] = pre_computed_findings

        print(
            f"[ExecSummary] ✅ OK — "
            f"framing_paragraph_len={len(result.get('framing_paragraph', ''))}, "
            f"key_findings_rows={len(result.get('key_findings', []))}, "
            f"core_message_len={len(result.get('core_message', ''))}"
        )
        return result

    except Exception as e:
        print(f"[ExecSummary] ❌ Error: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        raise


def _is_valid_table(t: dict) -> bool:
    """Validate that a table dict has required structure with non-empty rows."""
    return (isinstance(t, dict)
            and isinstance(t.get('rows'), list)
            and len(t.get('rows', [])) > 0
            and len(t.get('columns', [])) > 0)


def generate_methodology_section(state: PipelineState, api_key: str) -> Dict[str, Any]:
    """
    Generate the methodology section.

    Returns a dict with sources_table, classification_method, and analyzed_fields.
    """
    try:
        total_cases = state.total_cases
        date_range = convert_month_year_to_arabic(state.month_year) or "يناير — مارس 2026"
        guidebook_pages = getattr(state, 'guidebook_pages', 160)
        guidebook_faq_count = getattr(state, 'guidebook_faq_count', 25)
        guidebook_year = getattr(state, 'guidebook_year', '2025')

        # HARDCODED SECTION 2.2: منهجية التصنيف
        classification_method_hardcoded = (
            "يعتمد التحليل على شجرة قرار من أربعة مستويات، حيث طبيعة المطلوب وليس الصياغة هي المعيار الفاصل. يُطرح على كل حالة اختباران متتاليان:\n"
            "(1) الاختبار الأول: هل يُعبّر النص عن استياء، أو إبلاغ عن إخفاق، أو رغبة في تقديم بلاغ رسمي أو اعتراض؟\n"
            "(2) الاختبار الثاني: هل يطلب النص تنفيذ إجراء محدد كتقديم خدمة أو متابعة طلب أو تعديل بيانات، وليس مجرد الحصول على معلومة؟\n"
            "قواعد التصنيف:\n"
            "→ إن كان الجواب نعم على الأول: شكوى (حتى لو تضمّن النص طلباً لاتخاذ إجراء)؛\n"
            "→ إن كان الجواب نعم على الثاني فقط: طلب (دون أن يكون الغرض إبلاغاً عن مشكلة)؛\n"
            "→ إن كان الغرض سؤالاً أو استعلاماً عن معلومة: استفسار؛\n"
            "→ إن كان التعبير عن رضا وثناء: شكر وثناء."
        )

        # Build sources table
        sources_rows = [
            {
                "المصدر": "تحليل الشكاوى",
                "الطبيعة": "بيانات CRM — نصوص غير مهيكلة (تفاصيل الشكوى، الحلول، أسماء الخدمات، وصف الحالة)",
                "الحجم": f"{total_cases} حالة",
                "الفترة": date_range
            },
            {
                "المصدر": "دليل خدمات العملاء",
                "الطبيعة": "يغطي الخدمات المرورية والأمن والشهادات والتصاريح، مع التحقق من الأسئلة الشائعة وتحليل الفجوات",
                "الحجم": f"{guidebook_pages} صفحة، {guidebook_faq_count} أسئلة مصدقة",
                "الفترة": guidebook_year
            },
            {
                "المصدر": "منهجية إدارة الشكاوى",
                "الطبيعة": "وثيقة رسمية تحدد أنواع الشكاوى ومعايير المعالجة وKPIs",
                "الحجم": "6 صفحات",
                "الفترة": "-"
            }
        ]

        # HARDCODED SECTION 2.3: الحقول المحللة
        analyzed_fields_text = (
            "الحقول المنظمة المستخدمة في التحليل: رقم الطلب، الخدمة الرئيسية، نوع المكالمة الأصلي، حالة الطلب، قناة التواصل، والإدارة المختصة. "
            "الحقول غير المنظمة — محور هذا التحليل: تفاصيل الشكوى ورد المعالجة. هذان الحقلان يكشفان الطبيعة الحقيقية لكل حالة بعيداً عن التصنيف الشكلي في النظام."
        )

        return {
            "sources_table": sources_rows,
            "classification_method": classification_method_hardcoded,
            "analyzed_fields": analyzed_fields_text
        }

    except Exception as e:
        print(f"[Methodology] ❌ Error: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        raise


def run_stage6(
    state: PipelineState,
    excel_path: str,
    word_path: str,
    language: str = 'ar',
    api_key: str = ""
) -> PipelineState:
    """
    Stage 6: Artifact generator for complaints flow.

    Input: state with all_classified, analysis results, and report_sections
    Output: Excel and Word files on disk; report dictionary in state.report_json

    Args:
        state: Pipeline state from stages 1-5
        excel_path: Path to save Excel workbook
        word_path: Path to save Word document
        language: 'ar' or 'en'
        api_key: Anthropic API key for Word report generation

    Returns:
        Updated state with report_json dictionary
    """
    # Validation
    assert len(state.all_classified) == state.total_cases, "Case count mismatch"
    assert all(c.actual_contact_type for c in state.all_classified), "Missing contact types"
    assert all(0.0 <= c.confidence <= 1.0 for c in state.all_classified), "Invalid confidence scores"

    # FIX 1: Compute reclassification stats once — used by all downstream outputs
    reclassified = [
        c for c in state.all_classified
        if c.actual_contact_type != c.case_type
    ]
    state.reclassified_count = len(reclassified)
    state.reclassification_rate = (
        state.reclassified_count / state.total_cases * 100
        if state.total_cases > 0 else 0.0
    )

    # CRITICAL: Always clear and regenerate report sections to ensure LLM prompts
    # use the latest reconciled data (state.journey_map, state.gap_table, etc.)
    # and updated friction counts. Stale cached sections will have outdated prose.
    print(f"[Stage6] Clearing cached report_sections_ar and report_sections_en to ensure fresh generation...")
    state.report_sections_ar = {}
    state.report_sections_en = {}

    print(f"[Stage6] Calling _generate_report_sections with fresh state...")
    _generate_report_sections(state, api_key)
    print(f"[Stage6] After generation: {len(state.report_sections_ar or {})} sections in report_sections_ar")

    # Generate Excel
    generate_excel(state, excel_path)

    # Generate report dictionary first so state.report_json is always populated
    state.report_json = generate_json_report(state)

    # Translate Arabic JSON to English using the same LLM as the rest of the pipeline
    state.report_json_en = translate_report_to_english(state.report_json, api_key)
    if state.report_json_en:
        print("[Stage6] English translation of report JSON complete.")
    else:
        print("[Stage6] WARNING: English translation failed or was skipped.")

    # Generate Word report (skipped gracefully when word_path is None)
    generate_word_report(state, word_path, language, api_key)

    return state
