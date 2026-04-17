"""
STAGE 6: Artifact Generator

Generates:
1. Excel workbook (openpyxl) — 10 sheets with classification results
2. Word report (sword-word-builder) — 9-section bilingual report

Uses state.report_sections from Stage 4/5 to build Word document.
"""

import json
import anthropic
from pathlib import Path
from typing import Dict, Any, List
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .state import PipelineState, CaseRow

try:
    from sword_word_builder import WordBuilder, DocumentConfig, TextStyle, TableStyle, CoverPage
except ImportError:
    WordBuilder = None


# Excel formatting
HEADER_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_ROW_FILL = PatternFill(start_color="F8F9F9", end_color="F8F9F9", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Sheet names in order
SHEET_NAMES = [
    'ملخص',                    # Summary
    'كل الحالات',              # All Cases
    'طلبات',                   # Service Requests
    'استفسارات',               # Information Inquiries
    'متابعات',                 # Status Follow-ups
    'مشاكل جهات أخرى',         # Cross-Entity
    'استفسارات الموقع',        # Location Inquiries
    'بلاغات تقنية',            # Tech Incidents
    'استفسارات مالية',         # Financial Inquiries
    'إعادة التصنيف',           # Misclassified Cases
]


def generate_excel(state: PipelineState, output_path: str) -> None:
    """Generate Excel workbook with classification results."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Add summary sheet
    ws_summary = wb.create_sheet('ملخص', 0)
    _populate_summary_sheet(ws_summary, state)

    # Add all cases sheet
    ws_all = wb.create_sheet('كل الحالات', 1)
    _populate_all_cases_sheet(ws_all, state.all_classified)

    # Add category-specific sheets
    wb.create_sheet('طلبات', 2)
    wb.create_sheet('استفسارات', 3)
    wb.create_sheet('متابعات', 4)
    wb.create_sheet('مشاكل جهات أخرى', 5)
    wb.create_sheet('استفسارات الموقع', 6)
    wb.create_sheet('بلاغات تقنية', 7)
    wb.create_sheet('استفسارات مالية', 8)

    # Add misclassified cases sheet
    ws_misclass = wb.create_sheet('إعادة التصنيف', 9)
    misclassified = [c for c in state.all_classified if c.misclassification != 'OK']
    _populate_all_cases_sheet(ws_misclass, misclassified)

    # Color misclassified tab
    ws_misclass.sheet_properties.tabColor = "C0392B"

    # Save
    wb.save(output_path)


def _populate_summary_sheet(ws, state: PipelineState) -> None:
    """Populate summary sheet with metrics."""
    ws.merge_cells('A1:E1')
    ws['A1'] = f"ملخص تحليل استفسارات شرطة الفجيرة — {state.month_year or 'Q1 2026'}"
    ws['A1'].font = Font(bold=True, size=14)

    ws.merge_cells('A2:E2')
    ws['A2'] = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d')}"

    # Total cases
    row = 4
    ws[f'A{row}'] = "1. إجمالي الحالات"
    ws[f'A{row}'].font = Font(bold=True)

    row = 5
    ws[f'A{row}'] = "إجمالي الحالات المعالجة"
    ws[f'B{row}'] = len(state.all_classified)
    ws[f'C{row}'] = "100.0%"

    # Distribution by contact type
    row = 7
    ws[f'A{row}'] = "2. توزيع حسب نوع التصنيف"
    ws[f'A{row}'].font = Font(bold=True)

    contact_types = {}
    for case in state.all_classified:
        ct = case.actual_contact_type
        contact_types[ct] = contact_types.get(ct, 0) + 1

    row = 8
    total = len(state.all_classified)
    for contact_type, count in sorted(contact_types.items(), key=lambda x: x[1], reverse=True):
        ws[f'A{row}'] = contact_type
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"{(count/total*100):.1f}%"
        row += 1


def _populate_all_cases_sheet(ws, cases: List[CaseRow]) -> None:
    """Populate sheet with case data."""
    headers = [
        'رقم الطلب',
        'تفاصيل الطلب',
        'الحل',
        'الخدمة',
        'الخدمة الرئيسية',
        'نوع المكالمة',
        'التصنيف الفعلي',
        'السبب',
        'إعادة التصنيف',
        'قناة التقديم',
        'حالة SLA',
        'تاريخ الإنشاء',
        'الإدارة',
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Write data
    for row_idx, case in enumerate(cases, 2):
        data = [
            case.case_number,
            case.case_title[:50],  # Truncate for display
            case.resolution_response[:50],
            case.service_name,
            case.service_name,
            case.case_type,
            case.actual_contact_type,
            case.classification_reason,
            case.misclassification,
            case.case_channel,
            case.sla_color,
            case.date_opened,
            '',  # Admin
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = BORDER

            # Alternate row coloring
            if row_idx % 2 == 0:
                cell.fill = ALT_ROW_FILL

            # RTL alignment for Arabic
            cell.alignment = Alignment(horizontal='right' if col_idx <= 8 else 'left', vertical='top', wrap_text=True)

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 60
    for col in range(4, 14):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Freeze header
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(cases) + 1}'


def generate_word_report(
    state: PipelineState,
    output_path: str,
    language: str = 'ar',
    api_key: str = ""
) -> None:
    """
    Generate Word report using sword-word-builder.

    Args:
        state: Pipeline state with report_sections
        output_path: Path to save .docx
        language: 'ar' or 'en'
        api_key: Anthropic API key for LLM report generation
    """
    if WordBuilder is None:
        raise ImportError("sword-word-builder not installed")

    # Build report content via LLM if not in state
    if not state.report_sections:
        _generate_report_sections(state, api_key)

    # Create builder
    is_arabic = language == 'ar'
    builder = WordBuilder(
        DocumentConfig(
            default_font="Calibri",
            default_rtl=is_arabic,
            heading_font="Arial",
            accent_color="B68A35",  # Gold
            secondary_color="1A6080",  # Dark blue
            body_color="333333",
            line_spacing=14,
        )
    )

    # Add cover page
    if is_arabic:
        title = "نبض الفجيرة"
        subtitle = f"تقرير تحليل الاستفسارات — {state.month_year or 'الربع الأول 2026'}"
    else:
        title = "Fujairah Pulse"
        subtitle = f"Inquiry Analysis Report — {state.month_year or 'Q1 2026'}"

    cover = CoverPage.preset(
        title=title,
        subtitle=subtitle,
        metadata={
            "Total Cases": state.total_cases,
            "Report Date": state.created_at.split('T')[0],
        }
    )
    builder.add_cover_page(cover)

    # Add sections
    section_order = [
        'executive_summary',
        'methodology',
        'classification_summary',
        'workload_distribution',
        'top_patterns',
        'friction_analysis',
        'gap_analysis',
        'faq_summary',
        'recommendations',
        'conclusion',
    ]

    for section_key in section_order:
        if section_key not in state.report_sections:
            continue

        section_data = state.report_sections[section_key]

        # Get language-specific content
        if is_arabic:
            heading = section_data.get('heading_ar', section_data.get('heading', ''))
            body = section_data.get('body_ar', section_data.get('body', ''))
        else:
            heading = section_data.get('heading_en', section_data.get('heading', ''))
            body = section_data.get('body_en', section_data.get('body', ''))

        if not heading:
            continue

        # Add section heading
        builder.add_heading(heading, level=2, rtl=is_arabic)

        # Add body text
        if body:
            builder.add_paragraph(body, rtl=is_arabic)

        # Add tables if present
        if 'tables' in section_data:
            tables = section_data['tables']
            if isinstance(tables, list) and tables:
                for table_data in tables:
                    if isinstance(table_data, (list, dict)):
                        try:
                            builder.add_table(table_data, rtl=is_arabic)
                        except Exception as e:
                            print(f"Warning: Failed to add table in {section_key}: {e}")

        # Add spacer between sections
        builder.add_spacer(12)

    # Save
    builder.save(output_path)


def _generate_report_sections(state: PipelineState, api_key: str = "") -> None:
    """Generate report sections via LLM."""
    if not api_key:
        # Fallback to basic structure if no API key
        _create_basic_report_sections(state)
        return

    client = anthropic.Anthropic(api_key=api_key)

    # Build context from state
    summary_text = _build_summary_context(state)

    prompt = f"""Generate a comprehensive Arabic-English bilingual report structure for an inquiry analysis.

Context:
{summary_text}

Generate report sections with the following structure. Return as JSON with this format:
{{
  "section_slug": {{
    "heading": "English Heading",
    "heading_ar": "العنوان بالعربية",
    "body": "English body text...",
    "body_ar": "النص العربي...",
    "tables": [
      {{"column1": "value1", "column2": "value2"}},
      ...
    ]
  }}
}}

Required sections:
1. executive_summary - Key findings and metrics
2. methodology - Analysis approach and data sources
3. workload_distribution - Distribution of inquiry types
4. top_patterns - Most common patterns identified
5. friction_analysis - Customer journey friction points
6. gap_analysis - Information gaps identified
7. faq_summary - Frequently asked questions
8. recommendations - Actionable recommendations
9. conclusion - Summary and next steps

Make the content specific to the data provided. Use bilingual output (Arabic and English) for all content.
Include relevant statistics and percentages."""

    try:
        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4000,
            messages=[
                {
                    "role": "user",
                    "content": prompt
                }
            ]
        )

        # Extract JSON from response
        response_text = message.content[0].text

        # Try to parse JSON from response
        import re
        json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
        if json_match:
            try:
                sections = json.loads(json_match.group())
                state.report_sections = sections
                return
            except json.JSONDecodeError:
                pass

        # Fallback to basic structure if parsing fails
        _create_basic_report_sections(state)

    except Exception as e:
        print(f"Warning: Failed to generate report sections via LLM: {e}")
        _create_basic_report_sections(state)


def _build_summary_context(state: PipelineState) -> str:
    """Build summary context for report generation."""
    lines = [
        f"Total Cases: {state.total_cases}",
        f"Date: {state.month_year or 'Q1 2026'}",
    ]

    # Add distribution
    if state.all_classified:
        type_counts = {}
        for case in state.all_classified:
            ct = case.actual_contact_type
            type_counts[ct] = type_counts.get(ct, 0) + 1

        lines.append("\nClassification Distribution:")
        for ct, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
            pct = count / state.total_cases * 100
            lines.append(f"  {ct}: {count} ({pct:.1f}%)")

    # Add patterns
    if state.patterns:
        lines.append(f"\nIdentified Patterns: {len(state.patterns)} clusters")
        for pattern in state.patterns[:5]:
            lines.append(f"  - {pattern.sub_theme}: {pattern.case_count} cases")

    # Add friction points
    if state.journey_map:
        lines.append(f"\nFriction Points: {len(state.journey_map)}")
        for friction in state.journey_map[:5]:
            lines.append(f"  - {friction.friction_point}: {friction.case_count} cases")

    # Add gaps
    if state.gap_table:
        critical_gaps = [g for g in state.gap_table if g.severity == 'Critical']
        lines.append(f"\nGaps Identified: {len(state.gap_table)} total, {len(critical_gaps)} critical")

    # Add FAQs
    if state.validated_faqs:
        lines.append(f"\nValidated FAQs: {len(state.validated_faqs)}")

    return "\n".join(lines)


def _create_basic_report_sections(state: PipelineState) -> None:
    """Create basic report structure without LLM."""
    state.report_sections = {
        'executive_summary': {
            'heading': 'Executive Summary',
            'heading_ar': 'الملخص التنفيذي',
            'body': f'This report analyzes {state.total_cases} customer inquiries and complaints. Key findings include identification of customer service patterns, friction points in the customer journey, and recommendations for service improvement.',
            'body_ar': f'يحلل هذا التقرير {state.total_cases} استفسار وشكوى من المتعاملين. تشمل النتائج الرئيسية تحديد أنماط الخدمة، والنقاط الصعبة في رحلة المتعامل، والتوصيات لتحسين الخدمة.',
        },
        'methodology': {
            'heading': 'Methodology',
            'heading_ar': 'المنهجية',
            'body': 'Analysis was conducted using a six-stage pipeline: (1) Schema validation, (2) Rule-based classification, (3) LLM-enhanced classification, (4) Pattern analysis, (5) Gap identification, and (6) Report generation.',
            'body_ar': 'تم إجراء التحليل باستخدام خط أنابيب متقدم يضم ستة مراحل: (1) التحقق من الصيغة، (2) التصنيف القائم على القواعد، (3) التصنيف المحسّن بالذكاء الاصطناعي، (4) تحليل الأنماط، (5) تحديد الفجوات، و (6) توليد التقرير.',
        },
        'classification_summary': {
            'heading': 'Classification Summary',
            'heading_ar': 'ملخص التصنيف',
            'body': 'Cases were classified into eight categories based on customer intent and interaction patterns.',
            'body_ar': 'تم تصنيف الحالات في ثماني فئات بناءً على نوايا المتعامل وأنماط التفاعل.',
        },
        'recommendations': {
            'heading': 'Recommendations',
            'heading_ar': 'التوصيات',
            'body': 'Based on the analysis, the following recommendations are provided for improving service delivery and customer satisfaction.',
            'body_ar': 'بناءً على التحليل، يتم تقديم التوصيات التالية لتحسين جودة الخدمة ورضا المتعاملين.',
        },
    }


def run_stage6(
    state: PipelineState,
    excel_path: str,
    word_path: str,
    language: str = 'ar',
    api_key: str = ""
) -> PipelineState:
    """
    Stage 6: Artifact generator.

    Input: state with all_classified and analysis results
    Output: Excel and Word files on disk
    """
    # Validation
    assert len(state.all_classified) == state.total_cases, "Case count mismatch"
    assert all(c.actual_contact_type for c in state.all_classified), "Missing contact types"
    assert all(0.0 <= c.confidence <= 1.0 for c in state.all_classified), "Invalid confidence scores"

    # Generate Excel
    generate_excel(state, excel_path)

    # Generate Word report
    generate_word_report(state, word_path, language, api_key)

    return state
