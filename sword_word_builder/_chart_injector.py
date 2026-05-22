"""
Native Word chart injection for sword-word-builder.

Charts are two-pass:
  Pass 1 – during add_chart(), a sentinel bookmark paragraph is inserted.
  Pass 2 – during save()/build(), the .docx ZIP is patched to replace
            each sentinel with a proper DrawingML inline drawing and the
            corresponding chart XML + embedded Excel workbook.

Nothing in this module is part of the public API.
"""
from __future__ import annotations

import io
import zipfile
import re
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from lxml import etree

from .config import ChartStyle

# ---------------------------------------------------------------------------
# OOXML namespace constants
# ---------------------------------------------------------------------------
_NS = {
    "w":   "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "r":   "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "a":   "http://schemas.openxmlformats.org/drawingml/2006/main",
    "c":   "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "wp":  "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing",
    "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
    "ct":  "http://schemas.openxmlformats.org/package/2006/content-types",
}

_CM_TO_EMU = 360000  # 1 cm = 360 000 EMU

_CHART_CT   = "application/vnd.openxmlformats-officedocument.drawingml.chart+xml"
_XLSX_CT    = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_CHART_REL  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart"
_PKG_REL    = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/package"

# Sentinel text embedded as a w:bookmarkStart name in the placeholder paragraph.
# Format: __swb_chart_<index>__
SENTINEL_PREFIX = "__swb_chart_"
SENTINEL_SUFFIX = "__"


def sentinel_name(idx: int) -> str:
    return f"{SENTINEL_PREFIX}{idx}{SENTINEL_SUFFIX}"


# ---------------------------------------------------------------------------
# Pending chart descriptor
# ---------------------------------------------------------------------------

@dataclass
class _PendingChart:
    index: int
    data: dict          # {title, categories, series: [{name, values, color?}]}
    chart_type: str     # "column" | "bar" | "line"
    style: ChartStyle


# ---------------------------------------------------------------------------
# Excel workbook builder
# ---------------------------------------------------------------------------

def _build_xlsx(data: dict) -> bytes:
    """Build an embedded Excel workbook for the chart data.

    Layout (matches reference implementation):
      Row 1:  [empty] | series1_name | series2_name | ...
      Row 2:  cat1    | val1_s1      | val1_s2      | ...
      Row N+1: catN   | valN_s1      | valN_s2      | ...
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    categories = data.get("categories", [])
    series_list = data.get("series", [])

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    alt_fill = PatternFill("solid", fgColor="DCE6F1")
    plain_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: empty A1, then series names in columns B, C, ...
    ws.cell(1, 1, "")
    for ci, ser in enumerate(series_list, start=2):
        c = ws.cell(1, ci, ser.get("name", f"Series {ci - 1}"))
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    # Rows 2+: category in col A, values per series in cols B, C, ...
    for ri, cat in enumerate(categories, start=2):
        fill = alt_fill if ri % 2 == 0 else plain_fill
        cc = ws.cell(ri, 1, str(cat))
        cc.font = Font(bold=True)
        cc.fill = fill
        cc.border = border
        for ci, ser in enumerate(series_list, start=2):
            vals = ser.get("values", [])
            vi = ri - 2  # 0-based index into values
            val = vals[vi] if vi < len(vals) else None
            vc = ws.cell(ri, ci, val if val is not None else "")
            vc.fill = fill
            vc.border = border
            if isinstance(val, (int, float)):
                vc.alignment = Alignment(horizontal="right")

    # Column widths
    ws.column_dimensions["A"].width = 12
    for ci in range(2, len(series_list) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Chart XML builder
# ---------------------------------------------------------------------------

def _build_chart_xml(chart: _PendingChart) -> bytes:
    """Construct DrawingML chartSpace XML for a column, bar, or line chart."""
    data = chart.data
    style = chart.style
    categories = data.get("categories", [])
    series_list = data.get("series", [])
    n_cats = len(categories)
    title_text = data.get("title") or ""

    # Default colors cycling
    default_colors = style.series_colors
    if not default_colors:
        default_colors = ["4472C4", "ED7D31", "A9D18E", "FFC000"]

    def _c(tag: str) -> str:
        return f"{{{_NS['c']}}}{tag}"

    def _a(tag: str) -> str:
        return f"{{{_NS['a']}}}{tag}"

    root = etree.Element(_c("chartSpace"), nsmap={
        "c": _NS["c"],
        "a": _NS["a"],
        "r": _NS["r"],
    })

    etree.SubElement(root, _c("date1904")).set("val", "0")
    etree.SubElement(root, _c("lang")).set("val", "en-US")
    etree.SubElement(root, _c("roundedCorners")).set("val", "0")

    chart_el = etree.SubElement(root, _c("chart"))

    # Title
    if title_text:
        title_el = etree.SubElement(chart_el, _c("title"))
        tx = etree.SubElement(title_el, _c("tx"))
        rich = etree.SubElement(tx, _c("rich"))
        etree.SubElement(rich, _a("bodyPr"))
        etree.SubElement(rich, _a("lstStyle"))
        p = etree.SubElement(rich, _a("p"))
        r = etree.SubElement(p, _a("r"))
        if style.title_font_size:
            rPr = etree.SubElement(r, _a("rPr"))
            rPr.set("lang", "ar-SA")
            rPr.set("sz", str(style.title_font_size * 100))
        etree.SubElement(r, _a("t")).text = title_text
        etree.SubElement(title_el, _c("overlay")).set("val", "0")
    else:
        etree.SubElement(chart_el, _c("autoTitleDeleted")).set("val", "1")

    plot_area = etree.SubElement(chart_el, _c("plotArea"))
    etree.SubElement(plot_area, _c("layout"))

    # Choose chart element
    ctype = chart.chart_type
    if ctype == "pie":
        chart_body = etree.SubElement(plot_area, _c("pieChart"))
        etree.SubElement(chart_body, _c("varyColors")).set("val", "1")
    elif ctype == "line":
        chart_body = etree.SubElement(plot_area, _c("lineChart"))
        etree.SubElement(chart_body, _c("grouping")).set("val", "standard")
        etree.SubElement(chart_body, _c("varyColors")).set("val", "0")
    else:
        chart_body = etree.SubElement(plot_area, _c("barChart"))
        bar_dir = "bar" if ctype == "bar" else "col"
        etree.SubElement(chart_body, _c("barDir")).set("val", bar_dir)
        etree.SubElement(chart_body, _c("grouping")).set("val", "clustered")
        etree.SubElement(chart_body, _c("varyColors")).set("val", "0")

    if ctype == "pie":
        # Pie chart: one series, colors assigned per slice via <c:dPt>
        ser = series_list[0] if series_list else {}
        ser_name = ser.get("name", "Series 1")
        values = ser.get("values", [])

        ser_el = etree.SubElement(chart_body, _c("ser"))
        etree.SubElement(ser_el, _c("idx")).set("val", "0")
        etree.SubElement(ser_el, _c("order")).set("val", "0")

        # Series name from cell B1
        tx = etree.SubElement(ser_el, _c("tx"))
        strRef = etree.SubElement(tx, _c("strRef"))
        etree.SubElement(strRef, _c("f")).text = "Sheet1!$B$1"
        strCache = etree.SubElement(strRef, _c("strCache"))
        etree.SubElement(strCache, _c("ptCount")).set("val", "1")
        pt = etree.SubElement(strCache, _c("pt"))
        pt.set("idx", "0")
        etree.SubElement(pt, _c("v")).text = ser_name

        # One <c:dPt> per slice with individual color
        for pt_idx in range(n_cats):
            slice_color = default_colors[pt_idx % len(default_colors)]
            dpt = etree.SubElement(ser_el, _c("dPt"))
            etree.SubElement(dpt, _c("idx")).set("val", str(pt_idx))
            dpt_spPr = etree.SubElement(dpt, _c("spPr"))
            dpt_fill = etree.SubElement(dpt_spPr, _a("solidFill"))
            etree.SubElement(dpt_fill, _a("srgbClr")).set("val", slice_color)

        # Categories: column A, rows 2 to n_cats+1
        cat = etree.SubElement(ser_el, _c("cat"))
        cat_strRef = etree.SubElement(cat, _c("strRef"))
        etree.SubElement(cat_strRef, _c("f")).text = f"Sheet1!$A$2:$A${n_cats + 1}"
        cat_cache = etree.SubElement(cat_strRef, _c("strCache"))
        etree.SubElement(cat_cache, _c("ptCount")).set("val", str(n_cats))
        for ci, cat_val in enumerate(categories):
            cpt = etree.SubElement(cat_cache, _c("pt"))
            cpt.set("idx", str(ci))
            etree.SubElement(cpt, _c("v")).text = str(cat_val)

        # Values: column B, rows 2 to n_cats+1
        val_el = etree.SubElement(ser_el, _c("val"))
        numRef = etree.SubElement(val_el, _c("numRef"))
        etree.SubElement(numRef, _c("f")).text = f"Sheet1!$B$2:$B${n_cats + 1}"
        numCache = etree.SubElement(numRef, _c("numCache"))
        etree.SubElement(numCache, _c("formatCode")).text = "General"
        etree.SubElement(numCache, _c("ptCount")).set("val", str(n_cats))
        for vi, v in enumerate(values):
            if vi >= n_cats:
                break
            if v is None:
                continue
            vpt = etree.SubElement(numCache, _c("pt"))
            vpt.set("idx", str(vi))
            etree.SubElement(vpt, _c("v")).text = str(v)

        # Data labels for pie charts: show values on slices
        if style.show_data_labels:
            dLbls = etree.SubElement(ser_el, _c("dLbls"))
            etree.SubElement(dLbls, _c("showVal")).set("val", "1")
            etree.SubElement(dLbls, _c("showCatName")).set("val", "0")
            etree.SubElement(dLbls, _c("showSerName")).set("val", "0")
            etree.SubElement(dLbls, _c("showPercent")).set("val", "0")
            etree.SubElement(dLbls, _c("dLblPos")).set("val", "bestFit")

    else:
        # Series — reference layout: series in columns (B, C, ...), categories in rows (A2:AN+1)
        for ser_idx, ser in enumerate(series_list):
            raw_color = ser.get("color")
            # color may be a single hex string or a list of hex strings (one per bar/point)
            if isinstance(raw_color, list):
                per_point_colors = [c.lstrip("#").upper() for c in raw_color]
                ser_color = per_point_colors[0]  # series-level fallback = first color
            else:
                per_point_colors = None
                ser_color = (raw_color or default_colors[ser_idx % len(default_colors)]).lstrip("#").upper()

            ser_name = ser.get("name", f"Series {ser_idx + 1}")
            values = ser.get("values", [])
            col_letter = get_column_letter(ser_idx + 2)  # B, C, D, ...

            ser_el = etree.SubElement(chart_body, _c("ser"))
            etree.SubElement(ser_el, _c("idx")).set("val", str(ser_idx))
            etree.SubElement(ser_el, _c("order")).set("val", str(ser_idx))

            # Series name: row 1, column for this series
            tx = etree.SubElement(ser_el, _c("tx"))
            strRef = etree.SubElement(tx, _c("strRef"))
            etree.SubElement(strRef, _c("f")).text = f"Sheet1!${col_letter}$1"
            strCache = etree.SubElement(strRef, _c("strCache"))
            etree.SubElement(strCache, _c("ptCount")).set("val", "1")
            pt = etree.SubElement(strCache, _c("pt"))
            pt.set("idx", "0")
            etree.SubElement(pt, _c("v")).text = ser_name

            # Per-point color overrides via <c:dPt> (one per bar/point)
            if per_point_colors:
                for pt_idx in range(len(values)):
                    pt_color = per_point_colors[pt_idx % len(per_point_colors)]
                    dpt = etree.SubElement(ser_el, _c("dPt"))
                    etree.SubElement(dpt, _c("idx")).set("val", str(pt_idx))
                    etree.SubElement(dpt, _c("invertIfNegative")).set("val", "0")
                    dpt_spPr = etree.SubElement(dpt, _c("spPr"))
                    dpt_fill = etree.SubElement(dpt_spPr, _a("solidFill"))
                    etree.SubElement(dpt_fill, _a("srgbClr")).set("val", pt_color)
                    dpt_ln = etree.SubElement(dpt_spPr, _a("ln"))
                    etree.SubElement(etree.SubElement(dpt_ln, _a("solidFill")), _a("srgbClr")).set("val", pt_color)

            # Shape properties — series-level fill (used when no dPt override)
            spPr = etree.SubElement(ser_el, _c("spPr"))
            solidFill = etree.SubElement(spPr, _a("solidFill"))
            srgbClr = etree.SubElement(solidFill, _a("srgbClr"))
            srgbClr.set("val", ser_color)
            ln = etree.SubElement(spPr, _a("ln"))
            if ctype == "line":
                lineFill = etree.SubElement(ln, _a("solidFill"))
                lineClr = etree.SubElement(lineFill, _a("srgbClr"))
                lineClr.set("val", ser_color)
                ln.set("w", "25400")  # ~2pt
            else:
                solidFill2 = etree.SubElement(ln, _a("solidFill"))
                etree.SubElement(solidFill2, _a("srgbClr")).set("val", ser_color)

            # Marker for line charts
            if ctype == "line":
                marker = etree.SubElement(ser_el, _c("marker"))
                etree.SubElement(marker, _c("symbol")).set("val", "circle")
                etree.SubElement(marker, _c("size")).set("val", "5")
                mSpPr = etree.SubElement(marker, _c("spPr"))
                mFill = etree.SubElement(mSpPr, _a("solidFill"))
                mClr = etree.SubElement(mFill, _a("srgbClr"))
                mClr.set("val", ser_color)

            # Categories: column A, rows 2 to n_cats+1
            cat = etree.SubElement(ser_el, _c("cat"))
            cat_strRef = etree.SubElement(cat, _c("strRef"))
            etree.SubElement(cat_strRef, _c("f")).text = f"Sheet1!$A$2:$A${n_cats + 1}"
            cat_cache = etree.SubElement(cat_strRef, _c("strCache"))
            etree.SubElement(cat_cache, _c("ptCount")).set("val", str(n_cats))
            for ci, cat_val in enumerate(categories):
                cpt = etree.SubElement(cat_cache, _c("pt"))
                cpt.set("idx", str(ci))
                etree.SubElement(cpt, _c("v")).text = str(cat_val)

            # Values: same column, rows 2 to n_cats+1
            val_el = etree.SubElement(ser_el, _c("val"))
            numRef = etree.SubElement(val_el, _c("numRef"))
            etree.SubElement(numRef, _c("f")).text = f"Sheet1!${col_letter}$2:${col_letter}${n_cats + 1}"
            numCache = etree.SubElement(numRef, _c("numCache"))
            etree.SubElement(numCache, _c("formatCode")).text = "General"
            etree.SubElement(numCache, _c("ptCount")).set("val", str(n_cats))
            for vi, v in enumerate(values):
                if vi >= n_cats:
                    break
                if v is None:
                    continue  # omit None — preserves index, shows gap
                vpt = etree.SubElement(numCache, _c("pt"))
                vpt.set("idx", str(vi))
                etree.SubElement(vpt, _c("v")).text = str(v)

            if ctype == "line":
                etree.SubElement(ser_el, _c("smooth")).set("val", "0")

        # Axes (not used for pie charts)
        etree.SubElement(chart_body, _c("axId")).set("val", "1")
        etree.SubElement(chart_body, _c("axId")).set("val", "2")

        # Category axis
        cat_ax = etree.SubElement(plot_area, _c("catAx"))
        etree.SubElement(cat_ax, _c("axId")).set("val", "1")
        cat_scaling = etree.SubElement(cat_ax, _c("scaling"))
        etree.SubElement(cat_scaling, _c("orientation")).set("val", "minMax")
        etree.SubElement(cat_ax, _c("delete")).set("val", "0")
        etree.SubElement(cat_ax, _c("axPos")).set("val", "b")
        etree.SubElement(cat_ax, _c("majorTickMark")).set("val", "out")
        etree.SubElement(cat_ax, _c("minorTickMark")).set("val", "none")
        etree.SubElement(cat_ax, _c("tickLblPos")).set("val", "nextTo")
        etree.SubElement(cat_ax, _c("crossAx")).set("val", "2")
        etree.SubElement(cat_ax, _c("crosses")).set("val", "autoZero")
        etree.SubElement(cat_ax, _c("auto")).set("val", "1")
        etree.SubElement(cat_ax, _c("lblAlgn")).set("val", "ctr")
        etree.SubElement(cat_ax, _c("lblOffset")).set("val", "100")
        etree.SubElement(cat_ax, _c("noMultiLvlLbl")).set("val", "0")

        if style.axis_font_size:
            _add_axis_txPr(cat_ax, style.axis_font_size, _c, _a)

        # Optional x-axis title
        if style.x_axis_title:
            _add_axis_title(cat_ax, style.x_axis_title, _c, _a)

        # Value axis — empty <c:scaling/> matches reference (no orientation child)
        val_ax = etree.SubElement(plot_area, _c("valAx"))
        etree.SubElement(val_ax, _c("axId")).set("val", "2")
        etree.SubElement(val_ax, _c("scaling"))  # intentionally empty
        etree.SubElement(val_ax, _c("delete")).set("val", "0")
        etree.SubElement(val_ax, _c("axPos")).set("val", "l")
        if style.show_gridlines:
            etree.SubElement(val_ax, _c("majorGridlines"))
        etree.SubElement(val_ax, _c("majorTickMark")).set("val", "out")
        etree.SubElement(val_ax, _c("minorTickMark")).set("val", "none")
        etree.SubElement(val_ax, _c("tickLblPos")).set("val", "nextTo")
        etree.SubElement(val_ax, _c("crossAx")).set("val", "1")
        etree.SubElement(val_ax, _c("crosses")).set("val", "autoZero")

        if style.axis_font_size:
            _add_axis_txPr(val_ax, style.axis_font_size, _c, _a)

        # Optional y-axis title
        if style.y_axis_title:
            _add_axis_title(val_ax, style.y_axis_title, _c, _a)

    # Plot area background color
    bg_color = style.background_color.lstrip("#").upper()
    if bg_color:
        plot_spPr = etree.SubElement(plot_area, _c("spPr"))
        bg_fill = etree.SubElement(plot_spPr, _a("solidFill"))
        etree.SubElement(bg_fill, _a("srgbClr")).set("val", bg_color)

    etree.SubElement(chart_el, _c("dispBlanksAs")).set("val", "gap")

    # Legend
    if style.show_legend:
        legend = etree.SubElement(chart_el, _c("legend"))
        etree.SubElement(legend, _c("legendPos")).set("val", style.legend_position)
        etree.SubElement(legend, _c("overlay")).set("val", "0")
        if style.legend_font_size:
            _add_axis_txPr(legend, style.legend_font_size, _c, _a)

    etree.SubElement(chart_el, _c("plotVisOnly")).set("val", "1")

    # Chart-level text properties
    txpr = etree.SubElement(root, _c("txPr"))
    etree.SubElement(txpr, _a("bodyPr"))
    etree.SubElement(txpr, _a("lstStyle"))
    txpr_p = etree.SubElement(txpr, _a("p"))
    txpr_pPr = etree.SubElement(txpr_p, _a("pPr"))
    etree.SubElement(txpr_pPr, _a("defRPr")).set("sz", "1800")
    etree.SubElement(txpr_p, _a("endParaRPr")).set("lang", "en-US")

    # External data link (to embedded Excel)
    extData = etree.SubElement(root, _c("externalData"))
    extData.set(f"{{{_NS['r']}}}id", "rId1")
    etree.SubElement(extData, _c("autoUpdate")).set("val", "0")

    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _add_axis_txPr(parent_el, font_size_pt: int, c_tag, a_tag) -> None:
    """Inject a <c:txPr> (or <a:txPr> for legend) setting a specific font size."""
    txPr = etree.SubElement(parent_el, c_tag("txPr"))
    etree.SubElement(txPr, a_tag("bodyPr"))
    etree.SubElement(txPr, a_tag("lstStyle"))
    p = etree.SubElement(txPr, a_tag("p"))
    pPr = etree.SubElement(p, a_tag("pPr"))
    defRPr = etree.SubElement(pPr, a_tag("defRPr"))
    defRPr.set("sz", str(font_size_pt * 100))


def _add_axis_title(ax_el, text: str, c_tag, a_tag) -> None:
    title_el = etree.SubElement(ax_el, c_tag("title"))
    tx = etree.SubElement(title_el, c_tag("tx"))
    rich = etree.SubElement(tx, c_tag("rich"))
    etree.SubElement(rich, a_tag("bodyPr"))
    etree.SubElement(rich, a_tag("lstStyle"))
    p = etree.SubElement(rich, a_tag("p"))
    r = etree.SubElement(p, a_tag("r"))
    etree.SubElement(r, a_tag("t")).text = text
    etree.SubElement(title_el, c_tag("overlay")).set("val", "0")


def _build_chart_rels(chart_n: int) -> bytes:
    """Build word/charts/_rels/chartN.xml.rels linking to the embedded Excel."""
    target = f"../embeddings/chart{chart_n}_data.xlsx"
    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'<Relationship Id="rId1" Type="{_PKG_REL}" Target="{target}"/>'
        "</Relationships>"
    )
    return xml.encode("utf-8")


# ---------------------------------------------------------------------------
# Main injector
# ---------------------------------------------------------------------------

def inject_charts(docx_buffer: io.BytesIO, pending: list[_PendingChart],
                  doc_pr_id_start: int = 100) -> io.BytesIO:
    """Post-process a .docx buffer, replacing sentinel paragraphs with native charts.

    Returns a new BytesIO with the patched document.
    """
    if not pending:
        return docx_buffer

    docx_buffer.seek(0)
    original = zipfile.ZipFile(docx_buffer, "r")

    # Read files we need to patch
    doc_xml_bytes  = original.read("word/document.xml")
    rels_xml_bytes = original.read("word/_rels/document.xml.rels")
    ct_xml_bytes   = original.read("[Content_Types].xml")

    # Parse rels to find max rId integer
    rels_tree = etree.fromstring(rels_xml_bytes)
    existing_ids = []
    for rel in rels_tree.iter():
        rid = rel.get("Id", "")
        m = re.match(r"rId(\d+)", rid)
        if m:
            existing_ids.append(int(m.group(1)))
    next_rid = max(existing_ids, default=0) + 1

    # Assign rIds for each chart
    chart_rids: list[str] = []
    for _ in pending:
        chart_rids.append(f"rId{next_rid}")
        next_rid += 1

    # ---------------------------------------------------------------------------
    # Rebuild document.xml root with a, c, wp namespace declarations.
    # This is critical: if we insert drawing XML without these namespaces declared
    # at root level, lxml emits inline ns0:/ns1: prefixes that Word cannot parse.
    # ---------------------------------------------------------------------------
    NS_W  = _NS["w"]
    NS_WP = _NS["wp"]
    NS_A  = _NS["a"]
    NS_C  = _NS["c"]
    NS_R  = _NS["r"]

    doc_root = etree.fromstring(doc_xml_bytes)
    nsmap = dict(doc_root.nsmap)
    nsmap["a"]  = NS_A
    nsmap["c"]  = NS_C
    nsmap["wp"] = NS_WP
    new_doc_root = etree.Element(doc_root.tag, attrib=dict(doc_root.attrib), nsmap=nsmap)
    for child in doc_root:
        new_doc_root.append(child)

    body = new_doc_root.find(f"{{{NS_W}}}body")

    for chart in pending:
        sname  = sentinel_name(chart.index)
        chart_n = chart.index + 1  # 1-based file numbering
        r_id   = chart_rids[chart.index]
        width_emu  = int(chart.style.width_cm  * _CM_TO_EMU)
        height_emu = int(chart.style.height_cm * _CM_TO_EMU)
        doc_pr_id  = doc_pr_id_start + chart.index

        # Find sentinel paragraph: contains a w:bookmarkStart with this name
        sentinel_para = None
        for para in new_doc_root.iter(f"{{{NS_W}}}p"):
            for bk in para.iter(f"{{{NS_W}}}bookmarkStart"):
                if bk.get(f"{{{NS_W}}}name") == sname:
                    sentinel_para = para
                    break
            if sentinel_para is not None:
                break

        if sentinel_para is None:
            continue  # sentinel not found; skip

        # Build replacement paragraph as SubElements — no fromstring!
        # Using SubElement on the existing tree keeps namespace declarations at root.
        parent = sentinel_para.getparent()
        idx_in_parent = list(parent).index(sentinel_para)
        parent.remove(sentinel_para)

        new_para = etree.Element(f"{{{NS_W}}}p")
        pPr = etree.SubElement(new_para, f"{{{NS_W}}}pPr")
        jc  = etree.SubElement(pPr, f"{{{NS_W}}}jc")
        jc.set(f"{{{NS_W}}}val", "center")

        run = etree.SubElement(new_para, f"{{{NS_W}}}r")
        drw = etree.SubElement(run, f"{{{NS_W}}}drawing")

        inl = etree.SubElement(drw, f"{{{NS_WP}}}inline",
                               {"distT": "0", "distB": "0", "distL": "0", "distR": "0"})
        etree.SubElement(inl, f"{{{NS_WP}}}extent",
                         {"cx": str(width_emu), "cy": str(height_emu)})
        etree.SubElement(inl, f"{{{NS_WP}}}effectExtent",
                         {"l": "0", "t": "0", "r": "0", "b": "0"})
        etree.SubElement(inl, f"{{{NS_WP}}}docPr",
                         {"id": str(doc_pr_id), "name": f"Chart {chart_n}",
                          "descr": f"Chart {chart_n}"})
        cnv   = etree.SubElement(inl, f"{{{NS_WP}}}cNvGraphicFramePr")
        locks = etree.SubElement(cnv, f"{{{NS_A}}}graphicFrameLocks",
                                 {"noChangeAspect": "1"})
        gfx   = etree.SubElement(inl, f"{{{NS_A}}}graphic")
        gfxD  = etree.SubElement(gfx, f"{{{NS_A}}}graphicData", {"uri": NS_C})
        chrt  = etree.SubElement(gfxD, f"{{{NS_C}}}chart")
        chrt.set(f"{{{NS_R}}}id", r_id)

        parent.insert(idx_in_parent, new_para)

    patched_doc_xml = etree.tostring(new_doc_root, xml_declaration=True,
                                     encoding="UTF-8", standalone=True)

    # Patch rels
    for chart in pending:
        chart_n = chart.index + 1
        r_id = chart_rids[chart.index]
        rel = etree.SubElement(rels_tree, "Relationship")
        rel.set("Id", r_id)
        rel.set("Type", _CHART_REL)
        rel.set("Target", f"charts/chart{chart_n}.xml")
    patched_rels_xml = etree.tostring(rels_tree, xml_declaration=True,
                                      encoding="UTF-8", standalone=True)

    # Patch [Content_Types].xml
    ct_tree = etree.fromstring(ct_xml_bytes)
    CT_NS = _NS["ct"]
    # Add xlsx Default if missing
    has_xlsx_default = any(
        el.get("Extension") == "xlsx"
        for el in ct_tree.iter(f"{{{CT_NS}}}Default")
    )
    if not has_xlsx_default:
        default_el = etree.SubElement(ct_tree, f"{{{CT_NS}}}Default")
        default_el.set("Extension", "xlsx")
        default_el.set("ContentType", _XLSX_CT)

    for chart in pending:
        chart_n = chart.index + 1
        ov_chart = etree.SubElement(ct_tree, f"{{{CT_NS}}}Override")
        ov_chart.set("PartName", f"/word/charts/chart{chart_n}.xml")
        ov_chart.set("ContentType", _CHART_CT)
        ov_xlsx = etree.SubElement(ct_tree, f"{{{CT_NS}}}Override")
        ov_xlsx.set("PartName", f"/word/embeddings/chart{chart_n}_data.xlsx")
        ov_xlsx.set("ContentType", _XLSX_CT)

    patched_ct_xml = etree.tostring(ct_tree, xml_declaration=True,
                                    encoding="UTF-8", standalone=True)

    # Build new ZIP
    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        # Write [Content_Types].xml first (ECMA-376 requirement)
        zout.writestr("[Content_Types].xml", patched_ct_xml)

        for name in original.namelist():
            if name == "[Content_Types].xml":
                continue  # already written
            elif name == "word/document.xml":
                zout.writestr(name, patched_doc_xml)
            elif name == "word/_rels/document.xml.rels":
                zout.writestr(name, patched_rels_xml)
            else:
                zout.writestr(name, original.read(name))

        # Add chart files
        for chart in pending:
            chart_n = chart.index + 1
            chart_xml  = _build_chart_xml(chart)
            chart_rels = _build_chart_rels(chart_n)
            xlsx_bytes = _build_xlsx(chart.data)

            zout.writestr(f"word/charts/chart{chart_n}.xml",         chart_xml)
            zout.writestr(f"word/charts/_rels/chart{chart_n}.xml.rels", chart_rels)
            zout.writestr(f"word/embeddings/chart{chart_n}_data.xlsx",  xlsx_bytes)

    original.close()
    out_buf.seek(0)
    return out_buf
